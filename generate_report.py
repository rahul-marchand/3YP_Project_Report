"""
Generate LaTeX tables and figures for chapter 11 from financial_model.xlsx.

Workflow:
    1. Edit Assumptions tab in Excel OR edit build_model.py inputs
    2. Run: python3 build_model.py        (rebuilds the xlsx)
    3. Run: python3 generate_report.py    (regenerates figures/financial/*.tex)
    4. Recompile the thesis

All output files are written to figures/financial/ and are \\input{}-able
directly into chap_11_financial.tex.
"""

import os
import shutil
import subprocess
from openpyxl import load_workbook

MODEL   = "/home/rahul/thesis_template/financial_model.xlsx"
OUT_DIR = "/home/rahul/thesis_template/figures/financial"

os.makedirs(OUT_DIR, exist_ok=True)

YEAR_LABELS = ["Y1", "Y2", "Y3", "Y4", "Y5", "Y6", "Y7"]
N = len(YEAR_LABELS)

# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------
def years(ws, row, col_start=3, n=N):
    """Return list of values across year columns for a given row."""
    return [ws.cell(row, col_start + i).value or 0 for i in range(n)]

def fmt_gbp(v, dec=0):
    """Format number as GBP with parens for negatives."""
    if v is None: v = 0
    s = f"\\pounds{abs(v):,.{dec}f}"
    return f"({s})" if v < 0 else s

def fmt_k(v, dec=0):
    """Format number as £k with parens for negatives."""
    if v is None: v = 0
    s = f"\\pounds{abs(v)/1000:,.{dec}f}k"
    return f"({s})" if v < 0 else s

def fmt_m(v, dec=2):
    """Format number as £M."""
    if v is None: v = 0
    s = f"\\pounds{abs(v)/1e6:,.{dec}f}M"
    return f"({s})" if v < 0 else s

def fmt_num(v, dec=0):
    if v is None: v = 0
    return f"{v:,.{dec}f}"

def fmt_pct(v, dec=1):
    if v is None: v = 0
    return f"{v*100:.{dec}f}\\%"

def coords(ys, xs=None):
    """Build pgfplots coordinate string from a list of y values.
    If xs is None, uses YEAR_LABELS. Uses £k scale for large values."""
    xs = xs or YEAR_LABELS
    return " ".join(f"({x},{y/1000:.1f})" for x, y in zip(xs, ys))

def latex_escape(s):
    return s.replace("&", "\\&").replace("%", "\\%").replace("_", "\\_")


# ---------------------------------------------------------------------------
# LOAD BASE MODEL
# ---------------------------------------------------------------------------
print("Loading base model...")
wb = load_workbook(MODEL, data_only=True)
A  = wb["Assumptions"]
UE = wb["Unit economics"]
PL = wb["P&L"]
CF = wb["Cash flow"]
BS = wb["Balance sheet"]
SC = wb["Scenarios"]

# Extract key time series
new_units       = years(A, 15)            # total headsets sold/yr (now derived)
inst_buyers     = years(A, 28)            # institutional buyers/yr (schools + clubs)
dtc_buyers      = years(A, 29)            # DTC buyers/yr (consumer, no subscription)
headsets_per_inst = A.cell(26, 2).value   # scalar (3)
sub_price       = A.cell(27, 2).value     # scalar (£600)
inst_marketing  = years(A, 30)            # institutional channel marketing £/yr
dtc_marketing   = years(A, 31)            # DTC channel marketing £/yr
subs_eoy        = years(A, 17)
avg_subs        = years(A, 18)
cum_units       = years(A, 19)
blended_arpu    = years(A, 32)
unit_cost       = years(A, 49)
total_fte       = years(A, 62)
loaded_payroll  = years(A, 66)
people_cost     = years(A, 69)
rd_nonpay       = years(A, 72)
marketing       = years(A, 73)
travel          = years(A, 74)
prof_fees       = years(A, 75)
it_software     = years(A, 76)
facilities      = years(A, 80)
insurance       = years(A, 83)
ga_nonpay       = [t + p + i + f + n for t, p, i, f, n
                   in zip(travel, prof_fees, it_software, facilities, insurance)]
reg_clinical    = years(A, 88)
capex_per_yr    = years(A, 103)

# P&L lines
hw_rev          = years(PL, 9)
sub_rev         = years(PL, 10)
research_rev    = years(PL, 11)
acc_rev         = years(PL, 12)
warranty_rev    = years(PL, 13)
total_rev       = years(PL, 14)
total_cogs      = years(PL, 21)
gross_profit    = years(PL, 23)
gross_margin    = years(PL, 24)
total_rd        = years(PL, 29)
total_reg       = years(PL, 36)
total_sm        = years(PL, 43)
total_ga        = years(PL, 54)
total_opex      = years(PL, 56)
ebitda          = years(PL, 58)
da              = years(PL, 61)
ebit            = years(PL, 63)
interest_income = years(PL, 64)
corp_tax        = years(PL, 69)
rd_credit       = years(PL, 71)
net_tax         = years(PL, 72)
net_income      = years(PL, 75)

# Cash flow lines
op_cash_pre_wc  = years(CF, 13)
inv_change      = years(CF, 17)
ar_change       = years(CF, 19)
ap_change       = years(CF, 21)
def_rev_change  = years(CF, 23)
total_wc_change = years(CF, 24)
ncfo            = years(CF, 26)
capex_cf        = years(CF, 29)
equity          = years(CF, 33)
net_change      = years(CF, 37)
closing_cash    = years(CF, 39)
cum_unfunded    = years(CF, 42)

# Summary metrics from Cash flow tab
min_cash        = CF.cell(45, 2).value
min_cash_year   = CF.cell(46, 2).value
peak_trough     = CF.cell(47, 2).value
peak_trough_yr  = CF.cell(48, 2).value
total_equity    = CF.cell(49, 2).value
total_capex_cf  = CF.cell(51, 2).value
first_pos_ops   = CF.cell(52, 2).value
first_selffund  = CF.cell(53, 2).value

# Unit economics — Y7 steady state (column I, last year of the 7-year horizon)
ue_hw_contrib    = UE.cell(11, 9).value
ue_sub_contrib   = UE.cell(17, 9).value
ue_ltv_inst      = UE.cell(57, 9).value   # 3-yr LTV per institutional customer
ue_ltv_dtc       = UE.cell(58, 9).value   # LTV per DTC customer (= HW contrib)
ue_cac_inst      = UE.cell(51, 9).value   # CAC per institutional customer
ue_cac_dtc       = UE.cell(54, 9).value   # CAC per DTC customer
ue_ltv_cac_inst  = UE.cell(59, 9).value   # LTV:CAC institutional (3-yr)
ue_ltv_cac_dtc   = UE.cell(60, 9).value   # LTV:CAC DTC
ue_payback_inst  = UE.cell(61, 9).value   # Payback institutional (months)

# Tornado data (Scenarios tab section 5, rows 43-48 after Best/Worst added)
tornado_data = []
for r in range(43, 49):
    driver = SC.cell(r, 1).value
    dn = SC.cell(r, 2).value or 0
    up = SC.cell(r, 3).value or 0
    mag = SC.cell(r, 4).value or 0
    if driver:
        tornado_data.append((driver, dn, up, mag))


# ---------------------------------------------------------------------------
# RUN ALL 4 SCENARIOS  (for scenario comparison table and figure)
# ---------------------------------------------------------------------------
def run_scenario(selector):
    work = f"/tmp/scen_gen_{selector}.xlsx"
    shutil.copy(MODEL, work)
    w = load_workbook(work)
    w["Scenarios"]["B5"] = selector
    w.save(work)
    out_dir = f"/tmp/scen_gen_out_{selector}"
    os.makedirs(out_dir, exist_ok=True)
    subprocess.run(
        ["soffice", "--headless", "--calc", "--convert-to", "xlsx",
         "--outdir", out_dir, work],
        check=True, capture_output=True,
    )
    out = os.path.join(out_dir, os.path.basename(work))
    w2 = load_workbook(out, data_only=True)
    s = w2["Scenarios"]
    # Section 4 headlines now at rows 31-39 (after ASP/ARPU drivers added)
    result = {
        "name":    s.cell(31, 2).value,
        "revenue": s.cell(32, 2).value,
        "gp":      s.cell(33, 2).value,
        "ebitda":  s.cell(34, 2).value,
        "ni":      s.cell(35, 2).value,
        "cash":    s.cell(36, 2).value,
        "trough":  s.cell(37, 2).value,
        "break":   s.cell(38, 2).value,
        "equity":  s.cell(39, 2).value,
    }
    # Trajectories for line-chart overlays
    pl = w2["P&L"]
    result["ebitda_series"]       = years(pl, 58)
    result["revenue_series"]      = years(pl, 14)
    result["closing_cash_series"] = years(w2["Cash flow"], 39)
    os.remove(work)
    shutil.rmtree(out_dir)
    return result

print("Running scenarios (base / downside / upside / reg delay / best / worst)...")
SCENARIO_RESULTS = {
    "Base":     run_scenario(1),
    "Downside": run_scenario(2),
    "Upside":   run_scenario(3),
    "RegDelay": run_scenario(4),
    "Best":     run_scenario(5),
    "Worst":    run_scenario(6),
}
print("  done.")


# ---------------------------------------------------------------------------
# WRITERS
# ---------------------------------------------------------------------------
def write(name, content):
    path = os.path.join(OUT_DIR, name)
    with open(path, "w") as f:
        f.write(content)
    print(f"  wrote {path}")


# ===========================================================================
# TABLES
# ===========================================================================

# ---- Table 1: Year-by-year operational drivers ------------------------------
years_header = " & ".join(YEAR_LABELS)
# Pre-built bold-wrapped header (avoids confusing f-string brace nesting)
years_header_bf = " & ".join(f"\\textbf{{{y}}}" for y in YEAR_LABELS)

# Table 1 is grouped by the P&L block each driver feeds: Revenue → COGS →
# Operating expenses → Capex. Abbreviated group label sits in a narrow leftmost
# column (only on the first row of each group), saving the vertical space the
# previous multicolumn sub-header rows occupied while preserving the grouping.
groups = [
    ("Rev.", [
        ("Institutional buyers (schools + clubs)", "\\cref{chapter:ent-gtm}",
         [fmt_num(v) for v in inst_buyers]),
        ("DTC buyers (direct-to-consumer)",     "\\cref{chapter:ent-gtm}",
         [fmt_num(v) for v in dtc_buyers]),
        ("Total headsets sold (\\emph{derived})", "this section",
         [fmt_num(v) for v in new_units]),
        ("Hardware ASP (\\pounds)",             "\\cref{sec:ent-unit-economics}",
         [fmt_num(300) for _ in range(N)]),
        ("Subscription price (\\pounds/headset/yr)", "\\cref{sec:ent-pricing-strategy}",
         [fmt_num(sub_price) for _ in range(N)]),
        ("Annual churn (\\%)",                  "\\cref{sec:ent-unit-economics}",
         [fmt_pct(A.cell(16, 2).value or 0.10) for _ in range(N)]),
        ("Active headset subs (\\emph{derived})", "this section",
         [fmt_num(v) for v in subs_eoy]),
    ]),
    ("COGS", [
        ("Effective unit cost (\\pounds)",      "\\cref{tab:unit-cost-stack}",
         [fmt_num(v) for v in unit_cost]),
    ]),
    ("OpEx", [
        ("Total FTE",                           "\\cref{tab:ent-ops-headcount}",
         [f"{v:.1f}" for v in total_fte]),
        ("Loaded payroll (\\pounds k, \\emph{derived})", "\\cref{sec:ent-ops-salaries}",
         [fmt_num(v/1000) for v in loaded_payroll]),
        ("R\\&D non-payroll (\\pounds k)",      "dev tools, CAD/sim licences",
         [fmt_num(v/1000) for v in rd_nonpay]),
        ("Regulatory \\& clinical direct (\\pounds k)", "\\cref{chapter:ent-regulatory}",
         [fmt_num(v/1000) for v in reg_clinical]),
        ("Marketing — institutional (\\pounds k)", "\\cref{chapter:ent-gtm}",
         [fmt_num(v/1000) for v in inst_marketing]),
        ("Marketing — DTC (\\pounds k)",        "\\cref{chapter:ent-gtm}",
         [fmt_num(v/1000) for v in dtc_marketing]),
        ("G\\&A non-payroll (\\pounds k)",      "travel, IT, facilities, insurance",
         [fmt_num(v/1000) for v in ga_nonpay]),
    ]),
    ("Capex", [
        ("Capex (\\pounds k)",                  "\\cref{sec:ent-manufacturing-strategy}",
         [f"{v/1000:.1f}" for v in capex_per_yr]),
    ]),
]
body_lines = []
for i, (group_label, rows) in enumerate(groups):
    for j, (label, source, vals) in enumerate(rows):
        group_cell = f"\\textit{{{group_label}}}" if j == 0 else ""
        body_lines.append(
            "        " + group_cell + " & " + label + " & " + source + " & " + " & ".join(vals) + " \\\\"
        )
body = "\n".join(body_lines)
write("tab_assumptions.tex", f"""% Auto-generated by generate_report.py — do not edit manually
\\begin{{table}}[htbp]
    \\centering
    \\caption[Year-by-year operational drivers]{{Year-by-year drivers feeding each P\\&L block of the financial model. Drivers are grouped by the P\\&L line they feed (\\textit{{Revenue}}, \\textit{{COGS}}, \\textit{{Operating expenses}}, \\textit{{Capex}}). Rows marked \\emph{{derived}} are computed from the other rows and retained as audit anchors. Scalar modelling conventions are tabulated in \\cref{{tab:ent-fin-conventions}}.}}
    \\label{{tab:ent-fin-assumptions}}
    \\footnotesize
    \\begin{{tabular}}{{lll{"r"*N}}}
        \\toprule
         & \\textbf{{Driver}} & \\textbf{{Source}} & {years_header_bf} \\\\
        \\midrule
{body}
        \\bottomrule
    \\end{{tabular}}
\\end{{table}}
""")


# ---- Table 1b: Scalar financial-modelling conventions -----------------------
# Read scalars from the Assumptions sheet (column B = scalar slot)
def asc(row):
    return A.cell(row, 2).value

oncost      = asc(65)
returns     = asc(23)
sub_cogs    = asc(50)
ew_gm       = asc(44)
dio         = asc(110)
dso         = asc(111)
dpo         = asc(112)
ct_rate     = asc(116)
rd_eff      = asc(118)
rd_eng      = asc(119)
rd_reg      = asc(120)
rd_clin     = asc(121)
rd_trial    = asc(122)
rd_npp      = asc(123)
cash_int    = asc(124)
sales_comm  = asc(67)
founder_adj = A.cell(64, 3).value  # Y1 value of the founder cash array

conventions_groups = [
    ("Revenue / COGS", [
        ("Returns allowance",
            f"{returns*100:.0f}\\% of gross hardware revenue",
            "\\cref{sec:ent-ops-warranty}; DOA + trial returns"),
        ("Subscription COGS",
            f"{sub_cogs*100:.0f}\\% of subscription revenue",
            "cloud, support, data egress"),
        ("Extended-warranty gross margin",
            f"{ew_gm*100:.0f}\\%",
            "\\cref{sec:ent-ops-warranty}"),
        ("Sales commission",
            f"{sales_comm*100:.0f}\\% of new ARR from Y4 onwards",
            "\\cref{sec:ent-ops-org}"),
        ("Subscription billing",
            "Annually in advance",
            "drives deferred revenue"),
    ]),
    ("Payroll", [
        ("Founder partial cash",
            f"\\pounds{abs(founder_adj):,.0f} reduction Y1--Y2",
            "vs full salary; pre-Series A only"),
    ]),
    ("Working capital", [
        ("DIO / DSO / DPO",
            f"{dio:.0f} / {dso:.0f} / {dpo:.0f} days",
            "\\cref{sec:ent-inventory}; medtech default"),
    ]),
    ("Tax", [
        ("UK corporation tax",
            f"{ct_rate*100:.0f}\\% main rate",
            "full loss carry-forward"),
        ("SME R\\&D tax credit",
            f"{rd_eff*100:.1f}\\% effective rate",
            "post-April 2024 SME scheme"),
        ("R\\&D qualifying split",
            f"{rd_eng*100:.0f}/{rd_reg*100:.0f}/{rd_clin*100:.0f}/{rd_trial*100:.0f}/{rd_npp*100:.0f}\\%",
            "eng/reg/clin/trial/non-payroll; refund in year"),
    ]),
    ("Accounting", [
        ("Depreciation",
            "SL; 3y IT/bench, 5y tooling/fit-out",
            "\\cref{sec:ent-ops-facilities}"),
        ("Interest on cash",
            f"{cash_int*100:.0f}\\% p.a. on opening cash",
            "business savings rate"),
        ("Inflation",
            "nil; constant 2026 GBP",
            "modelling convention"),
        ("Foreign exchange",
            "frozen 2026 rates",
            "modelling convention"),
    ]),
]
conv_body_lines = []
for i, (group_label, conv_rows) in enumerate(conventions_groups):
    for j, (label, value, source) in enumerate(conv_rows):
        group_cell = f"\\textit{{{group_label}}}" if j == 0 else ""
        conv_body_lines.append(
            "        " + group_cell + " & " + label + " & " + value + " & " + source + " \\\\"
        )
conventions_body = "\n".join(conv_body_lines)
write("tab_conventions.tex", f"""% Auto-generated by generate_report.py — do not edit manually
\\begin{{table}}[htbp]
    \\centering
    \\caption[Scalar assumptions]{{Scalar assumptions used by the financial model, grouped by function. Year-varying drivers are tabulated separately in \\cref{{tab:ent-fin-assumptions}}.}}
    \\label{{tab:ent-fin-conventions}}
    \\footnotesize
    \\begin{{tabular}}{{llll}}
        \\toprule
         & \\textbf{{Assumption}} & \\textbf{{Value}} & \\textbf{{Source / note}} \\\\
        \\midrule
{conventions_body}
        \\bottomrule
    \\end{{tabular}}
\\end{{table}}
""")


# ---- Table 2: P&L summary ---------------------------------------------------
def row_k(label, series, bold=False, indent=""):
    cells = " & ".join(fmt_num(v/1000) if v >= 0 else f"({fmt_num(-v/1000)})" for v in series)
    line = f"{indent}{label} & {cells}"
    if bold:
        line = f"\\textbf{{{indent}{label}}} & " + " & ".join(
            f"\\textbf{{{fmt_num(v/1000) if v >= 0 else '(' + fmt_num(-v/1000) + ')'}}}" for v in series)
    return "        " + line + " \\\\"

pnl_rows = [
    row_k("Hardware revenue", hw_rev),
    row_k("Subscription revenue", sub_rev),
]
# Only include optional revenue streams if they have non-zero values
if any(v for v in research_rev):
    pnl_rows.append(row_k("Research licence revenue", research_rev))
if any(v for v in acc_rev):
    pnl_rows.append(row_k("Accessories revenue", acc_rev))
if any(v for v in warranty_rev):
    pnl_rows.append(row_k("Extended warranty revenue", warranty_rev))
pnl_rows += [
    row_k("\\textbf{Total revenue}", total_rev, bold=True),
    "        \\midrule",
    row_k("Total COGS", [-v for v in total_cogs]),
    row_k("\\textbf{Gross profit}", gross_profit, bold=True),
    "        \\midrule",
    row_k("Operating expenses", [-(r + g + s + a) for r, g, s, a in
                                 zip(total_rd, total_reg, total_sm, total_ga)]),
    row_k("\\textbf{EBITDA}", ebitda, bold=True),
    "        \\midrule",
    row_k("Depreciation \\& amortisation", [-v for v in da]),
    row_k("\\textbf{EBIT (operating profit/loss)}", ebit, bold=True),
    "        \\midrule",
    row_k("+ Interest income on cash", interest_income),
    row_k("Net tax / (R\\&D credit)", [-v for v in net_tax]),
    row_k("\\textbf{Net income}", net_income, bold=True),
]

write("tab_pnl.tex", f"""% Auto-generated by generate_report.py — do not edit manually
\\begin{{table}}[htbp]
    \\centering
    \\caption[7-year profit and loss projection]{{Seven-year profit and loss projection, base case. All P\\&L figures in \\pounds thousand; positive numbers are revenues or profits, parenthesised numbers are costs or losses.}}
    \\label{{tab:ent-fin-pnl}}
    \\footnotesize
    \\begin{{tabular}}{{l{"r"*N}}}
        \\toprule
        \\textbf{{Line item (\\pounds k)}} & {years_header_bf} \\\\
        \\midrule
{chr(10).join(pnl_rows)}
        \\bottomrule
    \\end{{tabular}}
\\end{{table}}
""")


# ---- Table 3: Cash flow summary --------------------------------------------
cf_rows = [
    row_k("EBITDA", ebitda),
    row_k("$-$ Tax / + R\\&D credit", [-v for v in net_tax]),
    row_k("+ Interest income on cash", interest_income),
    row_k("+ Working capital change (deferred-revenue tailwind)", total_wc_change),
    row_k("\\textbf{Net cash from operations}", ncfo, bold=True),
    "        \\midrule",
    row_k("Capex", capex_cf),
    row_k("Equity injections (illustrative; see \\cref{chapter:ent-funding})", equity),
    row_k("\\textbf{Net change in cash}", net_change, bold=True),
    "        \\midrule",
    row_k("Closing cash (end of year)", closing_cash),
    row_k("Cumulative unfunded position (operating reality)", cum_unfunded),
]

write("tab_cashflow.tex", f"""% Auto-generated by generate_report.py — do not edit manually
\\begin{{table}}[htbp]
    \\centering
    \\caption[Seven-year cash flow projection]{{Seven-year cash flow projection, base case. All figures in \\pounds thousand. The operating-cash bridge exposes the working-capital line directly: it is positive in every year, reflecting the deferred-revenue tailwind from annual-in-advance subscription billing. The cumulative unfunded position tracks operations and investing only (\\emph{{excluding}} financing); its lowest point is the gross funding requirement the venture must raise before reaching self-funding.}}
    \\label{{tab:ent-fin-cashflow}}
    \\footnotesize
    \\begin{{tabular}}{{l{"r"*N}}}
        \\toprule
        \\textbf{{Line item (\\pounds k)}} & {years_header_bf} \\\\
        \\midrule
{chr(10).join(cf_rows)}
        \\bottomrule
    \\end{{tabular}}
\\end{{table}}
""")


# ---- Table 4: Unit economics (by channel) -----------------------------------
ue_body = "\n".join([
    f"        Hardware contribution per headset       & \\multicolumn{{2}}{{r}}{{{fmt_gbp(ue_hw_contrib)}}} \\\\",
    f"        Subscription contribution per headset/yr & \\multicolumn{{2}}{{r}}{{{fmt_gbp(ue_sub_contrib)}}} \\\\",
    f"        \\midrule",
    f"        \\textbf{{Per-channel metric}} & \\textbf{{Institutional}} & \\textbf{{DTC}} \\\\",
    f"        \\midrule",
    f"        3-year LTV per customer                 & {fmt_gbp(ue_ltv_inst)} & {fmt_gbp(ue_ltv_dtc)} \\\\",
    f"        CAC per new customer                    & {fmt_gbp(ue_cac_inst)} & {fmt_gbp(ue_cac_dtc)} \\\\",
    f"        LTV : CAC ratio                         & {ue_ltv_cac_inst:.1f}$\\times$ & {ue_ltv_cac_dtc:.1f}$\\times$ \\\\",
    f"        Payback period (months)                 & {ue_payback_inst:.1f} & instant\\textsuperscript{{*}} \\\\",
])
write("tab_unit_economics.tex", f"""% Auto-generated by generate_report.py — do not edit manually
\\begin{{table}}[htbp]
    \\centering
    \\caption[Steady-state unit economics (Y7), by channel]{{Steady-state unit economics at Y7 (mature P3, commercial ramp complete), split by channel. Institutional LTV reflects three headsets per customer with per-headset subscriptions; DTC LTV is the one-off hardware contribution per parent purchase. CAC blends the channel's marketing spend with sales commission (institutional only). \\textsuperscript{{*}}DTC payback is at point of sale: the contribution exceeds CAC, so the purchase is self-funding without time-discounting.}}
    \\label{{tab:ent-fin-unit-econ}}
    \\footnotesize
    \\begin{{tabular}}{{lrr}}
        \\toprule
        \\textbf{{Per-headset metric}} & \\multicolumn{{2}}{{r}}{{\\textbf{{Value}}}} \\\\
        \\midrule
{ue_body}
        \\bottomrule
    \\end{{tabular}}
\\end{{table}}
""")


# ---- Table 5: Scenario comparison -------------------------------------------
def sc_row(label, key, fmt=fmt_gbp):
    cells = []
    for s in ["Base", "Downside", "Upside", "RegDelay"]:
        v = SCENARIO_RESULTS[s][key]
        if fmt == fmt_gbp:
            cells.append(fmt_gbp(v))
        else:
            cells.append(str(v))
    return "        " + label + " & " + " & ".join(cells) + " \\\\"

write("tab_scenarios.tex", f"""% Auto-generated by generate_report.py — do not edit manually
\\begin{{table}}[htbp]
    \\centering
    \\caption[Scenario comparison]{{Scenario comparison across four cases. \\textbf{{Base}} uses central assumptions. \\textbf{{Downside}} compounds all headwinds (slower adoption, higher BOM, higher churn, higher payroll, more marketing, and higher regulatory cost). \\textbf{{Upside}} reverses each lever. \\textbf{{Reg Delay}} is a single-issue stress where UKCA certification slips by one year, shifting the entire commercial ramp right. The peak unfunded trough is the binding funding requirement.}}
    \\label{{tab:ent-fin-scenarios}}
    \\footnotesize
    \\begin{{tabular}}{{lrrrr}}
        \\toprule
        \\textbf{{Metric}} & \\textbf{{Base}} & \\textbf{{Downside}} & \\textbf{{Upside}} & \\textbf{{Reg Delay}} \\\\
        \\midrule
{sc_row("Y7 total revenue", "revenue")}
{sc_row("Y7 gross profit", "gp")}
{sc_row("Y7 EBITDA", "ebitda")}
{sc_row("Y7 net income", "ni")}
{sc_row("Y7 closing cash", "cash")}
{sc_row("Peak unfunded cash trough", "trough")}
{sc_row("First year of positive operating cash", "break", fmt=str)}
        \\bottomrule
    \\end{{tabular}}
\\end{{table}}
""")


# ===========================================================================
# FIGURES (standalone TikZ/pgfplots → compiled to PDF via lualatex)
# ===========================================================================
#
# Each figure is written as a standalone .tex file (self-contained document)
# and compiled to PDF with lualatex. The chapter includes the PDF via
# \includegraphics{figures/financial/fig_*.pdf} in its own \begin{figure} block.

STANDALONE_HEAD = r"""% Auto-generated by generate_report.py — do not edit manually
% Preamble mirrors oxengthesis.cls so figures match the body text:
% Carlito as main font, Latin Modern Math via unicode-math.
\documentclass[border=6pt]{standalone}
\usepackage{fontspec}
\setmainfont{Carlito}
\usepackage{unicode-math}
\setmathfont{Latin Modern Math}
\usepackage{pgfplots}
\pgfplotsset{compat=1.18}
\usepgfplotslibrary{groupplots}
\usetikzlibrary{patterns,positioning}
\begin{document}
"""
STANDALONE_HEAD_CANVA = r"""% Auto-generated by generate_report.py — do not edit manually
% Cream-background slide variant — same content as the report version with
% canvabg pagecolor for use against the presentation theme.
\documentclass[border=8pt]{standalone}
\usepackage{fontspec}
\setmainfont{Carlito}
\usepackage{unicode-math}
\setmathfont{Latin Modern Math}
\usepackage{xcolor}
\usepackage{pgfplots}
\pgfplotsset{compat=1.18}
\usepgfplotslibrary{groupplots}
\usetikzlibrary{patterns,positioning}
\definecolor{canvabg}{RGB}{246,244,241}
\pagecolor{canvabg}
\begin{document}
"""
STANDALONE_TAIL = r"""
\end{document}
"""

def write_figure(name, tikz_body, cream=False):
    """Write a standalone TikZ .tex file and compile it to PDF with lualatex.
    If cream=True, uses the canvabg pagecolor for slide use."""
    head = STANDALONE_HEAD_CANVA if cream else STANDALONE_HEAD
    tex_path = os.path.join(OUT_DIR, f"{name}.tex")
    with open(tex_path, "w") as f:
        f.write(head + tikz_body + STANDALONE_TAIL)
    result = subprocess.run(
        ["lualatex", "-interaction=nonstopmode", "-halt-on-error",
         f"-output-directory={OUT_DIR}", tex_path],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        print(f"  FAIL  {name}.pdf")
        tail = result.stdout.splitlines()[-30:] if result.stdout else []
        for line in tail:
            print(f"    | {line}")
    else:
        print(f"  wrote {name}.pdf")

# Common pgfplots axis options reused across figures
PGF_COMMON = r"""            width=13cm,
            height=7cm,
            xlabel={Year},
            symbolic x coords={Y1,Y2,Y3,Y4,Y5,Y6,Y7},
            xtick=data,
            grid=major,
            grid style={line width=.1pt, draw=gray!30},
            major grid style={line width=.2pt,draw=gray!50},
            legend style={
                font=\footnotesize,
                at={(0.02,0.98)},
                anchor=north west,
                legend cell align={left},
            },
            tick label style={font=\footnotesize},
            label style={font=\small},
"""

def coord_list(series, scale=1000):
    return " ".join(f"({y},{v/scale:.1f})" for y, v in zip(YEAR_LABELS, series))


# ---- Figure 1: Revenue mix stacked bar --------------------------------------
write_figure("fig_revenue_mix", f"""\\begin{{tikzpicture}}
    \\begin{{axis}}[
{PGF_COMMON}        ylabel={{Revenue (\\pounds k)}},
        ybar stacked,
        bar width=16pt,
        ymin=0,
        legend pos=north west,
        legend columns=1,
    ]
    \\addplot+[ybar,fill=blue!60,draw=blue!80] coordinates {{{coord_list(hw_rev)}}};
    \\addplot+[ybar,fill=teal!60,draw=teal!80] coordinates {{{coord_list(sub_rev)}}};
    \\addplot+[ybar,fill=orange!60,draw=orange!80] coordinates {{{coord_list(research_rev)}}};
    \\addplot+[ybar,fill=purple!40,draw=purple!60] coordinates {{{coord_list(acc_rev)}}};
    \\addplot+[ybar,fill=gray!40,draw=gray!60] coordinates {{{coord_list(warranty_rev)}}};
    \\legend{{Hardware, Subscription, Research licences, Accessories, Extended warranty}}
    \\end{{axis}}
\\end{{tikzpicture}}
""")


# ---- Figure 2: EBITDA trajectory --------------------------------------------
ebitda_coords = coord_list(ebitda)
write_figure("fig_ebitda_trajectory", f"""\\begin{{tikzpicture}}
    \\begin{{axis}}[
{PGF_COMMON}        ylabel={{EBITDA (\\pounds k)}},
        ymin=-900, ymax=1200,
    ]
    % Phase boundary markers (P1 ends Y2, P2 ends Y4)
    \\draw[dashed,gray!50] (axis cs:Y2,-900) -- (axis cs:Y2,1200);
    \\draw[dashed,gray!50] (axis cs:Y4,-900) -- (axis cs:Y4,1200);
    \\node[anchor=north,font=\\scriptsize\\itshape,gray!80] at (axis cs:Y1,1180) {{Phase 1}};
    \\node[anchor=north,font=\\scriptsize\\itshape,gray!80] at (axis cs:Y3,1180) {{Phase 2}};
    \\node[anchor=north,font=\\scriptsize\\itshape,gray!80] at (axis cs:Y7,1180) {{Phase 3}};
    % EBITDA trajectory
    \\addplot[thick,blue,mark=*,mark size=2.5pt,mark options={{solid,fill=blue}}] coordinates {{{ebitda_coords}}};
    % Break-even reference line
    \\draw[dashed,gray,thin] (axis cs:Y1,0) -- (axis cs:Y7,0);
    \\node[anchor=west,font=\\footnotesize,gray] at (axis cs:Y1,80) {{break-even}};
    \\end{{axis}}
\\end{{tikzpicture}}
""")


# ---- Figure 2b: Operating expense composition stacked bar -------------------
write_figure("fig_opex_composition", f"""\\begin{{tikzpicture}}
    \\begin{{axis}}[
{PGF_COMMON}        ylabel={{Operating expense (\\pounds k)}},
        ybar stacked,
        bar width=16pt,
        ymin=0,
        legend pos=north west,
        legend columns=2,
    ]
    \\addplot+[ybar,fill=blue!60,draw=blue!80] coordinates {{{coord_list(total_rd)}}};
    \\addplot+[ybar,fill=teal!60,draw=teal!80] coordinates {{{coord_list(total_reg)}}};
    \\addplot+[ybar,fill=orange!60,draw=orange!80] coordinates {{{coord_list(total_sm)}}};
    \\addplot+[ybar,fill=gray!50,draw=gray!70] coordinates {{{coord_list(total_ga)}}};
    \\legend{{R\\&D, Reg \\& clinical, Sales \\& marketing, G\\&A}}
    \\end{{axis}}
\\end{{tikzpicture}}
""")


# ---- Figure 2c: Combined P&L overview (4-panel) -----------------------------
gm_pct_coords = " ".join(f"({y},{v*100:.1f})" for y, v in zip(YEAR_LABELS, gross_margin))

# Build revenue mix bar chart dynamically — only include series with non-zero values
revenue_mix_series = [
    ("Hardware",     "blue!60",   "blue!80",   hw_rev),
    ("Subscription", "teal!60",   "teal!80",   sub_rev),
    ("Research",     "orange!60", "orange!80", research_rev),
    ("Accessories",  "purple!40", "purple!60", acc_rev),
    ("Warranty",     "gray!40",   "gray!60",   warranty_rev),
]
revenue_mix_active = [(n, fc, dc, s) for (n, fc, dc, s) in revenue_mix_series if any(v for v in s)]
revenue_mix_addplots = "\n".join(
    f"    \\addplot+[ybar,fill={fc},draw={dc}] coordinates {{{coord_list(s)}}};"
    for (n, fc, dc, s) in revenue_mix_active
)
revenue_mix_legend = ", ".join(n for (n, _, _, _) in revenue_mix_active)
write_figure("fig_pnl_overview", f"""\\begin{{tikzpicture}}
    \\begin{{groupplot}}[
        group style={{
            group size=2 by 2,
            horizontal sep=2.0cm,
            vertical sep=2.0cm,
        }},
        width=7.5cm,
        height=5.2cm,
        symbolic x coords={{Y1,Y2,Y3,Y4,Y5,Y6,Y7}},
        xtick=data,
        xlabel={{Year}},
        grid=major,
        grid style={{line width=.1pt, draw=gray!30}},
        major grid style={{line width=.2pt, draw=gray!50}},
        tick label style={{font=\\small}},
        label style={{font=\\normalsize}},
        title style={{font=\\small\\bfseries, yshift=-2pt}},
        legend style={{
            font=\\scriptsize,
            cells={{anchor=west}},
            inner sep=2pt,
            row sep=-1pt,
        }},
    ]

    % --- Panel (a): Revenue mix ---
    \\nextgroupplot[
        title={{(a) Revenue mix (\\pounds k)}},
        ylabel={{Revenue}},
        ybar stacked,
        bar width=8pt,
        ymin=0,
        legend pos=north west,
        legend columns=1,
    ]
{revenue_mix_addplots}
    \\legend{{{revenue_mix_legend}}}

    % --- Panel (b): Gross margin % ---
    \\nextgroupplot[
        title={{(b) Gross margin (\\%)}},
        ylabel={{GM \\%}},
        ymin=30, ymax=80,
    ]
    \\addplot[thick,blue,mark=*,mark size=2pt,mark options={{solid,fill=blue}}] coordinates {{{gm_pct_coords}}};

    % --- Panel (c): OpEx composition ---
    \\nextgroupplot[
        title={{(c) OpEx composition (\\pounds k)}},
        ylabel={{OpEx}},
        ybar stacked,
        bar width=8pt,
        ymin=0,
        legend pos=north west,
        legend columns=1,
    ]
    \\addplot+[ybar,fill=blue!60,draw=blue!80] coordinates {{{coord_list(total_rd)}}};
    \\addplot+[ybar,fill=teal!60,draw=teal!80] coordinates {{{coord_list(total_reg)}}};
    \\addplot+[ybar,fill=orange!60,draw=orange!80] coordinates {{{coord_list(total_sm)}}};
    \\addplot+[ybar,fill=gray!50,draw=gray!70] coordinates {{{coord_list(total_ga)}}};
    \\legend{{R\\&D, Reg \\& clinical, S\\&M, G\\&A}}

    % --- Panel (d): EBITDA trajectory with phase markers ---
    \\nextgroupplot[
        title={{(d) EBITDA (\\pounds k)}},
        ylabel={{EBITDA}},
        ymin=-900, ymax=1200,
    ]
    \\draw[dashed,gray!50] (axis cs:Y2,-900) -- (axis cs:Y2,1200);
    \\draw[dashed,gray!50] (axis cs:Y4,-900) -- (axis cs:Y4,1200);
    \\node[anchor=north,font=\\scriptsize\\itshape,gray!80] at (axis cs:Y1,1180) {{P1}};
    \\node[anchor=north,font=\\scriptsize\\itshape,gray!80] at (axis cs:Y3,1180) {{P2}};
    \\node[anchor=north,font=\\scriptsize\\itshape,gray!80] at (axis cs:Y7,1180) {{P3}};
    \\addplot[thick,blue,mark=*,mark size=2pt,mark options={{solid,fill=blue}}] coordinates {{{coord_list(ebitda)}}};
    \\draw[dashed,gray,thin] (axis cs:Y1,0) -- (axis cs:Y7,0);
    \\node[anchor=west,font=\\scriptsize,gray] at (axis cs:Y1,80) {{break-even}};

    \\end{{groupplot}}
\\end{{tikzpicture}}
""")


# ---- Figure 3: Cash trajectory ---------------------------------------------
closing_coords = coord_list(closing_cash)
unfunded_coords = coord_list(cum_unfunded)
write_figure("fig_cash_trajectory", f"""\\begin{{tikzpicture}}
    \\begin{{axis}}[
{PGF_COMMON}        ylabel={{Cash position (\\pounds k)}},
        ymin=-4000, ymax=3000,
        legend pos=south west,
    ]
    \\addplot[thick,blue,mark=*,mark size=2pt] coordinates {{{closing_coords}}};
    \\addplot[thick,red!70,dashed,mark=square*,mark size=2pt] coordinates {{{unfunded_coords}}};
    \\draw[dotted,gray,thin] (axis cs:Y1,0) -- (axis cs:Y7,0);
    \\legend{{Closing cash (with equity), Cumulative unfunded position}}
    \\end{{axis}}
\\end{{tikzpicture}}
""")


# ---- Figure 3b: Combined cash overview (2-panel) ----------------------------
# Cumulative working-capital contribution (used in panel b)
cum_wc_running = []
_running = 0
for v in total_wc_change:
    _running += v
    cum_wc_running.append(_running)
wc_annual_coords = coord_list(total_wc_change)
wc_cum_coords    = coord_list(cum_wc_running)

write_figure("fig_cash_overview", f"""\\begin{{tikzpicture}}
    \\begin{{groupplot}}[
        group style={{
            group size=2 by 1,
            horizontal sep=2.0cm,
        }},
        width=7.5cm,
        height=5.5cm,
        symbolic x coords={{Y1,Y2,Y3,Y4,Y5,Y6,Y7}},
        xtick=data,
        xlabel={{Year}},
        grid=major,
        grid style={{line width=.1pt, draw=gray!30}},
        major grid style={{line width=.2pt, draw=gray!50}},
        tick label style={{font=\\small}},
        label style={{font=\\normalsize}},
        title style={{font=\\small\\bfseries, yshift=-2pt}},
        legend style={{
            font=\\scriptsize,
            cells={{anchor=west}},
            inner sep=2pt,
            row sep=-1pt,
        }},
    ]

    % --- Panel (a): Cash trajectory ---
    \\nextgroupplot[
        title={{(a) Cash position (\\pounds k)}},
        ylabel={{Cash}},
        ymin=-4000, ymax=3000,
        legend pos=south west,
    ]
    \\addplot[thick,blue,mark=*,mark size=2pt,mark options={{solid,fill=blue}}] coordinates {{{closing_coords}}};
    \\addlegendentry{{Closing (with equity)}}
    \\addplot[thick,red!70,dashed,mark=square*,mark size=2pt,mark options={{solid,fill=red!70}}] coordinates {{{unfunded_coords}}};
    \\addlegendentry{{Cumulative unfunded}}
    \\draw[dotted,gray,thin] (axis cs:Y1,0) -- (axis cs:Y7,0);

    % --- Panel (b): Working capital contribution ---
    \\nextgroupplot[
        title={{(b) Working capital contribution (\\pounds k)}},
        ylabel={{WC inflow}},
        ymin=0,
        legend pos=north west,
    ]
    \\addplot[ybar,bar width=10pt,fill=teal!50,draw=teal!70] coordinates {{{wc_annual_coords}}};
    \\addlegendentry{{Annual}}
    \\addplot[thick,blue,mark=*,mark size=2pt,mark options={{solid,fill=blue}}] coordinates {{{wc_cum_coords}}};
    \\addlegendentry{{Cumulative}}

    \\end{{groupplot}}
\\end{{tikzpicture}}
""")


# ---- Figure 4: Tornado sensitivity ------------------------------------------
tornado_sorted = sorted(tornado_data, key=lambda x: x[3])
drivers_tex = ",".join(f"{{{latex_escape(d[0])}}}" for d in tornado_sorted)
down_coords = " ".join(f"({d[1]/1000:.0f},{{{latex_escape(d[0])}}})" for d in tornado_sorted)
up_coords = " ".join(f"({d[2]/1000:.0f},{{{latex_escape(d[0])}}})" for d in tornado_sorted)

write_figure("fig_tornado", f"""\\begin{{tikzpicture}}
    \\begin{{axis}}[
        width=13cm,
        height=6.5cm,
        xbar,
        bar width=10pt,
        xlabel={{Y7 EBITDA impact (\\pounds k)}},
        symbolic y coords={{{drivers_tex}}},
        ytick=data,
        y axis line style={{opacity=0}},
        enlarge y limits=0.15,
        nodes near coords,
        nodes near coords style={{font=\\tiny}},
        nodes near coords align={{horizontal}},
        grid=major,
        grid style={{line width=.1pt, draw=gray!30}},
        legend style={{
            at={{(0.02,0.02)}}, anchor=south west, font=\\footnotesize,
            legend cell align={{left}},
        }},
        tick label style={{font=\\footnotesize}},
        label style={{font=\\small}},
    ]
    \\addplot+[xbar,fill=red!50,draw=red!70] coordinates {{{down_coords}}};
    \\addplot+[xbar,fill=green!50,draw=green!70] coordinates {{{up_coords}}};
    \\legend{{Adverse move, Favourable move}}
    \\end{{axis}}
\\end{{tikzpicture}}
""")


# ---- Refined sensitivity analysis ------------------------------------------
# Tornado bars are loaded from the workbook (Scenarios tab section 5) — single
# source of truth. Reg-delay is appended here from the actual scenario data
# because it is a discrete timing shock, not a linear approximation.
#
# The shock magnitude (in %) for each loaded driver is recorded here so the
# margin-of-safety chart can convert absolute £ impacts back into "% shock to
# break-even" values. Order matches the build_model.py tornado definition.
tornado_shock_pct = {
    "Adoption ramp ±20%":          20,
    "Blended ARPU ±15%":           15,
    "BOM unit cost +30%/-15%":     30,
    "Subscription churn 10%→15%":  50,  # 10% → 15% is +50% relative
    "Hardware ASP ±15%":           15,
    "Headcount/payroll ±10%":      10,
}

# tornado_data was loaded earlier from Scenarios rows 39-44.
# Append the regulatory-delay impact from the scenario run (not in workbook tornado).
reg_delay_impact = SCENARIO_RESULTS["RegDelay"]["ebitda"] - SCENARIO_RESULTS["Base"]["ebitda"]
tornado_refined = [
    (label, dn, up) for (label, dn, up, mag) in tornado_data
] + [("Regulatory delay +1 year", reg_delay_impact, 0)]

# Joint-failure bound from the compound-downside scenario (cited in prose)
compound_dn_ebitda = SCENARIO_RESULTS["Downside"]["ebitda"]
compound_dn_trough = SCENARIO_RESULTS["Downside"]["trough"]

# Sort smallest-first so largest sits at the top of the chart
tornado_refined.sort(key=lambda x: max(abs(x[1]), abs(x[2])))

# LaTeX-escape labels for the refined tornado (the workbook labels contain
# unicode characters like ± and → that pgfplots' symbolic coords don't accept).
def tex_label(label):
    return (label
            .replace("±", "$\\pm$")
            .replace("→", "$\\to$")
            .replace("%", "\\%"))

tr_drivers_tex = ",".join(f"{{{tex_label(d[0])}}}" for d in tornado_refined)
tr_down_coords = " ".join(f"({d[1]/1000:.0f},{{{tex_label(d[0])}}})" for d in tornado_refined)
tr_up_coords   = " ".join(f"({d[2]/1000:.0f},{{{tex_label(d[0])}}})" for d in tornado_refined)

# Margin-of-safety: % deviation in each lever required to wipe Y7 EBITDA to zero
# (linear approximation from the loaded tornado coefficients; reg delay excluded
# because it is a discrete timing shock, not a percentage shock).
base_y9 = ebitda[6]
def shock_to_zero_pct(impact, shock_pct, base):
    """Return absolute % shock required to drive base to zero, capped at 200."""
    coeff_per_pct = abs(impact) / shock_pct
    if coeff_per_pct == 0:
        return 200
    return min(200, abs(base / coeff_per_pct))

# Friendly labels for the margin-of-safety chart (no shock magnitudes in the label)
mos_label_for = {
    "Adoption ramp ±20%":          "Adoption ramp",
    "Blended ARPU ±15%":           "Blended ARPU",
    "BOM unit cost +30%/-15%":     "BOM unit cost",
    "Subscription churn 10%→15%":  "Subscription churn",
    "Hardware ASP ±15%":           "Hardware ASP",
    "Headcount/payroll ±10%":      "Headcount/payroll",
}
mos_levers = []
for (label, dn, up, mag) in tornado_data:
    shock_pct = tornado_shock_pct.get(label, 20)
    impact = max(abs(dn), abs(up))
    mos_levers.append(
        (mos_label_for.get(label, label), shock_to_zero_pct(impact, shock_pct, base_y9))
    )
mos_levers.sort(key=lambda x: x[1])  # tightest at top

mos_drivers_tex = ",".join(f"{{{d[0]}}}" for d in reversed(mos_levers))
mos_coords      = " ".join(f"({d[1]:.0f},{{{d[0]}}})" for d in reversed(mos_levers))
# Auto-fit the x-axis to the data: ceil to the next 10, with a small headroom for labels
mos_max_val     = max(d[1] for d in mos_levers) if mos_levers else 100
mos_xmax        = int(((mos_max_val * 1.25) // 10 + 1) * 10)


# ---- Figure 4b: Sensitivity overview (2-panel) -----------------------------
write_figure("fig_sensitivity_overview", f"""\\begin{{tikzpicture}}
    \\begin{{groupplot}}[
        group style={{
            group size=2 by 1,
            horizontal sep=2.8cm,
        }},
        width=8cm,
        height=6.2cm,
        tick label style={{font=\\small}},
        label style={{font=\\normalsize}},
        title style={{font=\\small\\bfseries, yshift=-2pt}},
        grid=major,
        grid style={{line width=.1pt, draw=gray!30}},
        major grid style={{line width=.2pt, draw=gray!50}},
    ]

    % --- Panel (a): Refined tornado ---
    \\nextgroupplot[
        title={{(a) Tornado: Y7 EBITDA impact (\\pounds k)}},
        xbar,
        bar width=8pt,
        xlabel={{EBITDA delta (\\pounds k)}},
        symbolic y coords={{{tr_drivers_tex}}},
        ytick=data,
        y axis line style={{opacity=0}},
        enlarge y limits=0.10,
        legend style={{
            at={{(0.02,0.02)}},
            anchor=south west,
            font=\\scriptsize,
            cells={{anchor=west}},
        }},
    ]
    \\addplot+[xbar,fill=red!55,draw=red!75] coordinates {{{tr_down_coords}}};
    \\addplot+[xbar,fill=teal!55,draw=teal!75] coordinates {{{tr_up_coords}}};
    \\legend{{Adverse, Favourable}}

    % --- Panel (b): Margin of safety ---
    \\nextgroupplot[
        title={{(b) Margin of safety: \\% shock to break-even}},
        xbar,
        bar width=8pt,
        xlabel={{\\% deviation from base}},
        symbolic y coords={{{mos_drivers_tex}}},
        ytick=data,
        y axis line style={{opacity=0}},
        enlarge y limits=0.10,
        nodes near coords,
        nodes near coords style={{font=\\scriptsize}},
        nodes near coords align={{horizontal}},
        xmin=0, xmax={mos_xmax},
    ]
    \\addplot+[xbar,fill=orange!55,draw=orange!75] coordinates {{{mos_coords}}};

    \\end{{groupplot}}
\\end{{tikzpicture}}
""")


# ---- Table: KPI dashboard ---------------------------------------------------
# Targets are derived as practical buffers above the linear failure thresholds
# computed for the margin-of-safety chart. The threshold column makes the
# derivation verifiable for the reader.

# Compute concrete failure thresholds from mos data
mos_dict = {label: pct for (label, pct) in mos_levers}
y9_arpu = blended_arpu[6]
y9_churn_base = (asc(16) or 0.10)

adoption_thresh_pct = 100 - mos_dict.get("Adoption ramp", 35)
arpu_thresh_gbp     = y9_arpu * (1 - mos_dict.get("Blended ARPU", 42) / 100)
churn_thresh_pct    = y9_churn_base * (1 + mos_dict.get("Subscription churn", 200) / 100) * 100
bom_thresh_pct      = 100 + mos_dict.get("BOM unit cost", 120)
hc_thresh_pct       = 100 + mos_dict.get("Headcount/payroll", 100)
asp_thresh_gbp      = 300 * (1 - mos_dict.get("Hardware ASP", 85) / 100)
# ASP threshold is unrealistic (below unit cost); the binding floor is the unit cost
asp_thresh_str      = f"$\\leq \\pounds {unit_cost[6]:.0f}$ (cost floor)"

kpi_rows = [
    # (kpi, source, target, failure threshold, cadence)
    ("Quarterly new unit shipments",
     "Adoption",
     "$\\geq 80$\\,\\% of plan",
     f"$\\leq {adoption_thresh_pct:.0f}$\\,\\% of plan",
     "Quarterly"),
    ("Cumulative active subscribers",
     "Adoption",
     "$\\geq 80$\\,\\% of plan",
     "(as above)",
     "Quarterly"),
    ("Realised hardware ASP",
     "Hardware ASP",
     "$\\geq \\pounds 300$/unit",
     asp_thresh_str,
     "Per shipment"),
    ("Blended ARPU per active customer",
     "ARPU / tier mix",
     "$\\geq \\pounds 290$/yr",
     f"$\\leq \\pounds {arpu_thresh_gbp:.0f}$/yr",
     "Annually"),
    ("Annual subscription churn",
     "Churn",
     "$\\leq 12$\\,\\%",
     f"$\\geq {churn_thresh_pct:.0f}$\\,\\%",
     "Annually"),
    ("Unit cost per device",
     "BOM cost",
     "$\\leq$ phase $+$5\\,\\%",
     f"$\\geq {bom_thresh_pct:.0f}$\\,\\% of base",
     "Per batch"),
    ("Monthly cash burn",
     "Headcount",
     "$\\leq$ plan $+$10\\,\\%",
     f"$\\geq {hc_thresh_pct:.0f}$\\,\\% of plan",
     "Monthly"),
    ("NB milestone slip",
     "Reg timing",
     "$\\leq$ 3 months",
     "1 year (scenario)",
     "Quarterly"),
]
kpi_body = "\n".join(
    "        " + " & ".join([k[0], k[2], k[3], k[4]]) + " \\\\"
    for k in kpi_rows
)
write("tab_kpis.tex", f"""% Auto-generated by generate_report.py — do not edit manually
\\begin{{table}}[htbp]
    \\centering
    \\caption[Operational KPI dashboard]{{Operational KPIs derived from the dominant tornado levers of \\cref{{fig:ent-fin-sensitivity}}. Each KPI shows both the operational target (set conservatively above the failure threshold) and the linear failure threshold itself (the value at which Y7 EBITDA would cross zero under a single-variable shock). These thresholds are necessary but not sufficient; the joint-failure bound is given in the prose below.}}
    \\label{{tab:ent-fin-kpis}}
    \\footnotesize
    \\begin{{tabular}}{{p{{4.2cm}}p{{3.2cm}}p{{3.4cm}}l}}
        \\toprule
        \\textbf{{KPI}} & \\textbf{{Target}} & \\textbf{{Failure threshold}} & \\textbf{{Cadence}} \\\\
        \\midrule
{kpi_body}
        \\bottomrule
    \\end{{tabular}}
\\end{{table}}
""")


# ---- Figure 5: Scenario comparison (two-panel bar chart) -------------------
scen_names = ["Base", "Downside", "Upside", "RegDelay"]
scen_labels = {"Base": "Base", "Downside": "Downside", "Upside": "Upside", "RegDelay": "Reg Delay"}
trough_coords = " ".join(
    f"({{{scen_labels[s]}}},{-SCENARIO_RESULTS[s]['trough']/1000:.0f})"
    for s in scen_names
)
ebitda_coords_sc = " ".join(
    f"({{{scen_labels[s]}}},{SCENARIO_RESULTS[s]['ebitda']/1000:.0f})"
    for s in scen_names
)
labels_sc = ",".join(f"{{{scen_labels[s]}}}" for s in scen_names)

write_figure("fig_scenarios", f"""\\begin{{tikzpicture}}
    \\begin{{axis}}[
        width=7cm,
        height=6cm,
        name=left,
        ybar,
        bar width=16pt,
        ylabel={{Peak funding need (\\pounds k)}},
        symbolic x coords={{{labels_sc}}},
        xtick=data,
        ymin=0,
        nodes near coords,
        nodes near coords style={{font=\\tiny}},
        grid=major,
        grid style={{line width=.1pt, draw=gray!30}},
        tick label style={{font=\\footnotesize}},
        label style={{font=\\small}},
        title={{(a) Gross funding requirement}},
        title style={{font=\\small}},
    ]
    \\addplot+[ybar,fill=red!40,draw=red!70] coordinates {{{trough_coords}}};
    \\end{{axis}}
    \\begin{{axis}}[
        width=7cm,
        height=6cm,
        at=(left.east), anchor=west, xshift=1.5cm,
        ybar,
        bar width=16pt,
        ylabel={{Y7 EBITDA (\\pounds k)}},
        symbolic x coords={{{labels_sc}}},
        xtick=data,
        nodes near coords,
        nodes near coords style={{font=\\tiny}},
        grid=major,
        grid style={{line width=.1pt, draw=gray!30}},
        tick label style={{font=\\footnotesize}},
        label style={{font=\\small}},
        title={{(b) Y7 EBITDA}},
        title style={{font=\\small}},
    ]
    \\addplot+[ybar,fill=blue!40,draw=blue!70] coordinates {{{ebitda_coords_sc}}};
    \\end{{axis}}
\\end{{tikzpicture}}
""")


# ---- Figures 6a/b/c: Best vs Base vs Worst trajectories (chap 13 §13.1) -----
# Three line charts showing the joint scenarios from chap_13_risk.tex against
# the base case: EBITDA, closing cash, and revenue. Cash is the binding metric
# (worst case is unfundable, not just unprofitable); revenue isolates the
# reg-delay shift visually.
def _round_axis(lo, hi, step=500):
    """Round axis bounds outward to a clean multiple of `step` (in £k)."""
    import math
    return math.floor(lo / step) * step, math.ceil(hi / step) * step

def _series_k(scen, key):
    """Series in £k for a given scenario and trajectory key."""
    return [v / 1000 for v in SCENARIO_RESULTS[scen][key]]

def _coords_k(scen, key):
    return " ".join(f"({y},{v:.1f})" for y, v in zip(YEAR_LABELS, _series_k(scen, key)))

# Axis bounds for the chap-13 best/base/worst comparison.
# RegDelay is intentionally excluded — its 12-month timing slip is already
# baked into the Worst scenario, so plotting both would double-count.
ebitda_all = (_series_k("Best", "ebitda_series") + _series_k("Base", "ebitda_series")
              + _series_k("Worst", "ebitda_series"))
ymin_e, ymax_e = _round_axis(min(ebitda_all) - 50, max(ebitda_all) + 50)

cash_all = (_series_k("Best", "closing_cash_series") + _series_k("Base", "closing_cash_series")
            + _series_k("Worst", "closing_cash_series"))
ymin_c, ymax_c = _round_axis(min(cash_all) - 200, max(cash_all) + 200)

rev_all = (_series_k("Best", "revenue_series") + _series_k("Base", "revenue_series")
           + _series_k("Worst", "revenue_series"))
ymin_r, ymax_r = _round_axis(0, max(rev_all) + 200)

# Per-scenario plot styling — used in every scenario panel for consistency.
SCEN_STYLES = {
    "Best":  ("teal!70!black", "triangle*", ""),
    "Base":  ("blue",          "*",         ""),
    "Worst": ("red!70!black",  "square*",   ",dashed"),
}
SCEN_ORDER  = ["Best", "Base", "Worst"]
SCEN_LABELS = {"Best": "Best case", "Base": "Base case", "Worst": "Worst case"}

def _addplots(key, mark_size="2.5pt", scenarios=None):
    """Render the \\addplot lines for a chosen subset of scenarios."""
    scenarios = scenarios or SCEN_ORDER
    lines = []
    for s in scenarios:
        col, mark, extra = SCEN_STYLES[s]
        lines.append(
            f"    \\addplot[thick,{col},mark={mark},mark size={mark_size}{extra}] "
            f"coordinates {{{_coords_k(s, key)}}};"
        )
    return "\n".join(lines)

def _legend(scenarios=None, short=False):
    scenarios = scenarios or SCEN_ORDER
    if short:
        return ", ".join(scenarios)
    return ", ".join(SCEN_LABELS[s] for s in scenarios)

_legend_full  = _legend()
_legend_short = _legend(short=True)

# Axis options shared across the singleton scenario figures, mirroring
# PGF_COMMON but with grid lines stripped per chap-13 styling preference.
PGF_NO_GRID = r"""            width=13cm,
            height=7cm,
            xlabel={Year},
            symbolic x coords={Y1,Y2,Y3,Y4,Y5,Y6,Y7},
            xtick=data,
            legend style={
                font=\footnotesize,
                at={(0.02,0.98)},
                anchor=north west,
                legend cell align={left},
            },
            tick label style={font=\footnotesize},
            label style={font=\small},
"""

# --- (a) EBITDA trajectory ---
write_figure("fig_scenarios_ebitda", f"""\\begin{{tikzpicture}}
    \\begin{{axis}}[
{PGF_NO_GRID}        ylabel={{EBITDA (\\pounds k)}},
        ymin={ymin_e}, ymax={ymax_e},
        legend pos=south east,
    ]
    \\draw[dashed,gray!50] (axis cs:Y2,{ymin_e}) -- (axis cs:Y2,{ymax_e});
    \\draw[dashed,gray!50] (axis cs:Y4,{ymin_e}) -- (axis cs:Y4,{ymax_e});
    \\node[anchor=north,font=\\scriptsize\\itshape,gray!80] at (axis cs:Y1,{ymax_e-20}) {{Phase 1}};
    \\node[anchor=north,font=\\scriptsize\\itshape,gray!80] at (axis cs:Y3,{ymax_e-20}) {{Phase 2}};
    \\node[anchor=north,font=\\scriptsize\\itshape,gray!80] at (axis cs:Y7,{ymax_e-20}) {{Phase 3}};
{_addplots("ebitda_series")}
    \\draw[dotted,gray,thin] (axis cs:Y1,0) -- (axis cs:Y7,0);
    \\legend{{{_legend_full}}}
    \\end{{axis}}
\\end{{tikzpicture}}
""")

# --- (b) Closing cash trajectory (the binding metric) ---
write_figure("fig_scenarios_cash", f"""\\begin{{tikzpicture}}
    \\begin{{axis}}[
{PGF_NO_GRID}        ylabel={{Closing cash (\\pounds k)}},
        ymin={ymin_c}, ymax={ymax_c},
        legend pos=south west,
    ]
{_addplots("closing_cash_series")}
    \\draw[dotted,gray,thin] (axis cs:Y1,0) -- (axis cs:Y7,0);
    \\legend{{{_legend_full}}}
    \\end{{axis}}
\\end{{tikzpicture}}
""")

# --- (c) Revenue trajectory ---
write_figure("fig_scenarios_revenue", f"""\\begin{{tikzpicture}}
    \\begin{{axis}}[
{PGF_NO_GRID}        ylabel={{Total revenue (\\pounds k)}},
        ymin={ymin_r}, ymax={ymax_r},
        legend pos=north west,
    ]
{_addplots("revenue_series")}
    \\legend{{{_legend_full}}}
    \\end{{axis}}
\\end{{tikzpicture}}
""")

# --- Combined 3-panel figure (chap 13 inclusion + slide variants) ---
def _combined_body(scenarios, axis_line_style="", legend_extra=""):
    """Build the combined 3-panel figure body for a chosen scenario subset.
    Grid lines are intentionally omitted across all variants."""
    legend_short = _legend(scenarios, short=True)
    return f"""\\begin{{tikzpicture}}
    \\begin{{groupplot}}[
        group style={{
            group size=3 by 1,
            horizontal sep=1.6cm,
        }},
        width=6.2cm,
        height=5.5cm,
        symbolic x coords={{Y1,Y2,Y3,Y4,Y5,Y6,Y7}},
        xtick=data,
        xlabel={{Year}},
{axis_line_style}        tick label style={{font=\\scriptsize}},
        label style={{font=\\small}},
        title style={{font=\\small\\bfseries, yshift=-2pt}},
        legend style={{
            font=\\scriptsize,
            cells={{anchor=west}},
            inner sep=2pt,
            row sep=-1pt,
{legend_extra}        }},
    ]

    % --- (a) EBITDA ---
    \\nextgroupplot[
        title={{(a) EBITDA (\\pounds k)}},
        ylabel={{EBITDA}},
        ymin={ymin_e}, ymax={ymax_e},
        legend pos=north west,
    ]
{_addplots("ebitda_series", mark_size="2pt", scenarios=scenarios)}
    \\draw[dotted,gray,thin] (axis cs:Y1,0) -- (axis cs:Y7,0);
    \\legend{{{legend_short}}}

    % --- (b) Closing cash (binding metric) ---
    \\nextgroupplot[
        title={{(b) Closing cash (\\pounds k)}},
        ylabel={{Cash}},
        ymin={ymin_c}, ymax={ymax_c},
    ]
{_addplots("closing_cash_series", mark_size="2pt", scenarios=scenarios)}
    \\draw[dotted,gray,thin] (axis cs:Y1,0) -- (axis cs:Y7,0);

    % --- (c) Revenue ---
    \\nextgroupplot[
        title={{(c) Revenue (\\pounds k)}},
        ylabel={{Revenue}},
        ymin={ymin_r}, ymax={ymax_r},
    ]
{_addplots("revenue_series", mark_size="2pt", scenarios=scenarios)}

    \\end{{groupplot}}
\\end{{tikzpicture}}
"""

CANVA_AXIS = ("        axis line style={line width=0.9pt, draw=gray!70},\n"
              "        tick style={line width=0.7pt, color=gray!70},\n")
CANVA_LEGEND = "            fill=canvabg,\n            draw=gray!50,\n"

# Report version (chap 13, white background, all three scenarios)
write_figure("fig_scenarios_bestworst",
             _combined_body(scenarios=["Best", "Base", "Worst"]))

# Slide version: best + base + worst (cream)
write_figure("fig_scenarios_bestworst_canva",
             _combined_body(scenarios=["Best", "Base", "Worst"],
                            axis_line_style=CANVA_AXIS,
                            legend_extra=CANVA_LEGEND),
             cream=True)

# Slide version: best + base only (cream) — for the lead-in upside narrative
write_figure("fig_scenarios_bestbase_canva",
             _combined_body(scenarios=["Best", "Base"],
                            axis_line_style=CANVA_AXIS,
                            legend_extra=CANVA_LEGEND),
             cream=True)

print()
print(f"All outputs written to {OUT_DIR}")
print("Include figures in chap 11 via \\includegraphics{figures/financial/fig_*.pdf}")
print("Include tables via \\input{figures/financial/tab_*.tex}")
