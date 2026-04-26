"""
Financial model build script for the mTBI Screening Tool venture (chap 11).

Re-run to regenerate financial_model.xlsx from scratch. All inputs live in the
Assumptions tab; every other tab is pure formulas referencing it.

Conventions:
  - Blue text  = hardcoded user input
  - Black text = formula
  - Green text = cross-tab link
  - Orange text + cream fill = [PROVISIONAL] placeholder pending another chapter
  - Bold      = section totals / sub-totals
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

OUT = "/home/rahul/thesis_template/financial_model.xlsx"

# ---------------------------------------------------------------------------
# STYLE CONSTANTS
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"

def f(color="000000", bold=False, italic=False, size=10):
    return Font(name=FONT_NAME, size=size, color=color, bold=bold, italic=italic)

font_default    = f()
font_input      = f(color="0000FF")
font_formula    = f()
font_link       = f(color="008000")
font_prov       = f(color="C65911", bold=True)
font_section    = f(color="FFFFFF", bold=True, size=11)
font_subheader  = f(bold=True)
font_total      = f(bold=True)
font_note       = f(color="595959", italic=True, size=9)
font_title      = f(bold=True, size=14, color="1F4E78")
font_year_label = f(bold=True, color="FFFFFF")

fill_section    = PatternFill("solid", start_color="1F4E78")
fill_year       = PatternFill("solid", start_color="305496")
fill_subsection = PatternFill("solid", start_color="D9E1F2")
fill_prov       = PatternFill("solid", start_color="FFF2CC")
fill_total      = PatternFill("solid", start_color="EDEDED")

GBP = '£#,##0;(£#,##0);"-"'
NUM = '#,##0;(#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'
INT = '0'
FTE = '0.0;(0.0);"-"'

YEAR_LABELS = ["Y1", "Y2", "Y3", "Y4", "Y5", "Y6", "Y7"]
YEAR_COLS = "CDEFGHI"   # cols 3..9
N_YEARS = len(YEAR_LABELS)
UNITS_COL = 12   # column L
NOTES_COL = 13   # column M
LAST_COL = 13    # for section header merges and total-row fills

# Scenarios tab — active multiplier cell references (section 3 of Scenarios)
S_ADOPTION  = "Scenarios!$B$18"
S_BOM       = "Scenarios!$B$19"
S_CHURN     = "Scenarios!$B$20"
S_PAYROLL   = "Scenarios!$B$21"
S_MARKETING = "Scenarios!$B$22"
S_REG       = "Scenarios!$B$23"
S_REG_DELAY = "Scenarios!$B$24"

FMT_TEXT = '@'


def delayed_units(col, A_prefix=""):
    """
    Formula expression for reg-delay-shifted total headsets in a given column.

    Commercial years shift right by S_REG_DELAY years; research years
    (periods 1–3) are unaffected. Years that get shifted off the end repeat
    the pre-launch Y3 value.
    """
    prefix = A_prefix
    return (
        f"INDEX({prefix}$C$15:{prefix}$K$15, 1, "
        f"MAX({col}5 - {S_REG_DELAY}, MIN({col}5, 3)))"
    )


def delayed_inst(col, A_prefix=""):
    """
    Reg-delay-shifted institutional buyers (Assumptions row 28).
    Drives subscription accumulation (parents excluded) and sales commission.
    """
    prefix = A_prefix
    return (
        f"INDEX({prefix}$C$28:{prefix}$K$28, 1, "
        f"MAX({col}5 - {S_REG_DELAY}, MIN({col}5, 3)))"
    )


# ---------------------------------------------------------------------------
# CELL HELPERS (closed over a sheet `ws`)
# ---------------------------------------------------------------------------
class Sheet:
    def __init__(self, ws):
        self.ws = ws

    def cell(self, row, col, value=None, font=None, fill=None, fmt=None, align=None, comment=None):
        c = self.ws.cell(row=row, column=col, value=value)
        if font: c.font = font
        if fill: c.fill = fill
        if fmt:  c.number_format = fmt
        if align: c.alignment = align
        if comment: c.comment = Comment(comment, "Model")
        return c

    def section(self, row, text):
        for col in range(1, LAST_COL + 1):
            self.ws.cell(row=row, column=col).fill = fill_section
        self.cell(row, 1, text, font=font_section, fill=fill_section)
        self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST_COL)
        self.ws.row_dimensions[row].height = 20

    def label(self, row, text, units="", source="", subheader=False, total=False):
        fnt = font_total if total else (font_subheader if subheader else font_default)
        fl  = fill_total if total else None
        # Escape leading formula-trigger characters in label text
        if text and text[0] in ("=", "+", "-", "@"):
            text = " " + text
        self.cell(row, 1, text, font=fnt, fill=fl)
        if total:
            for col in range(2, LAST_COL + 1):
                self.ws.cell(row=row, column=col).fill = fill_total
        if units:
            self.cell(row, UNITS_COL, units, font=font_note)
        if source:
            if source[0] in ("=", "+", "-", "@"):
                source = " " + source
            self.cell(row, NOTES_COL, source, font=font_note)

    def scalar(self, row, value, fmt=GBP, prov=False, comment=None):
        fnt = font_prov if prov else font_input
        fl  = fill_prov if prov else None
        return self.cell(row, 2, value, font=fnt, fill=fl, fmt=fmt, comment=comment)

    def years(self, row, values, fmt=GBP, prov=False):
        fnt = font_prov if prov else font_input
        fl  = fill_prov if prov else None
        for i, v in enumerate(values):
            self.cell(row, 3 + i, v, font=fnt, fill=fl, fmt=fmt)

    def year_formulas(self, row, build, fmt=GBP, total=False):
        """build(col_letter, col_idx) -> formula string."""
        fnt = font_total if total else font_formula
        fl  = fill_total if total else None
        for i, col in enumerate(YEAR_COLS):
            self.cell(row, 3 + i, build(col, i), font=fnt, fill=fl, fmt=fmt)


# ---------------------------------------------------------------------------
# ASSUMPTIONS TAB
# ---------------------------------------------------------------------------
def build_assumptions(wb):
    ws = wb.active
    ws.title = "Assumptions"
    s = Sheet(ws)

    # Column widths
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 13
    for col in YEAR_COLS:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 44

    # ---- Title ----
    s.cell(1, 1, "mTBI Screening Tool — Financial Model — Assumptions", font=font_title)
    ws.merge_cells("A1:M1")
    s.cell(2, 1,
           "Single source of truth. Every other tab references this one. "
           "Edit only blue (input) or orange (provisional) cells. "
           "Set unused tier/role/capex slots to 0 — formulas adapt automatically.",
           font=font_note)
    ws.merge_cells("A2:M2")

    # ---- Year header rows ----
    s.label(4, "Year")
    for i, y in enumerate(YEAR_LABELS):
        c = s.cell(4, 3 + i, y, font=font_year_label,
                   align=Alignment(horizontal="center"), fill=fill_year)

    s.label(5, "Period number", units="(driver)")
    s.years(5, list(range(1, N_YEARS + 1)), fmt=INT)
    s.label(6, "Calendar year")
    s.years(6, [2026 + i for i in range(N_YEARS)], fmt=INT)
    s.label(7, "Phase", units="(driver)",
            source="1 = research validation, 2 = first commercial, 3 = scale-up")
    s.years(7, [1, 1, 2, 2, 3, 3, 3], fmt=INT)

    # ---- Section 0: Sanity checks ----
    s.section(9, "0. SANITY CHECKS")
    s.label(10, "Total headsets = 3 × inst + parent",
            source="Should display OK across the row (row 15 = $B$26·row28 + row29)")
    s.year_formulas(10,
        lambda col, i: f'=IF({col}15=$B$26*{col}28+{col}29, "OK", "FAIL")',
        fmt=FMT_TEXT)
    s.label(11, "Cash never below buffer",
            source="From Cash flow tab — checks all 9 closing-cash years")
    s.cell(11, 2,
           '=IF(MIN(\'Cash flow\'!C39:I39) >= $B$127, "OK", "BREACH")',
           font=font_formula)
    s.label(12, "Balance sheet balances",
            source="From Balance sheet tab — sum of absolute differences")
    s.cell(12, 2,
           '=IF(SUMPRODUCT(--(ABS(\'Balance sheet\'!C34:K34)<1))=9, "OK", "FAIL")',
           font=font_formula)

    # ---- Section 1: Adoption ramp (derived from GTM channel mix in section 3) ----
    s.section(14, "1. ADOPTION RAMP   (derived from GTM channel mix, §3)")
    s.label(15, "Total headsets sold per year (derived)", units="headsets",
            source="= 3 × institutional buyers + parent buyers (rows 28–29)")
    s.year_formulas(15,
        lambda col, i: f"=$B$26*{col}28 + {col}29", fmt=NUM)
    s.label(16, "Annual subscription churn", units="% per year",
            source="ch.9 §9.4 indicative; applied to start-of-year base")
    s.scalar(16, 0.10, fmt=PCT)
    s.label(17, "Active headset subscriptions (end of year)", units="subs",
            source="prior × (1 − churn) + institutional buyers × headsets/inst × attach × adoption")
    s.year_formulas(17,
        lambda col, i: (
            f"={delayed_inst(col)} * $B$26 * {S_ADOPTION} * $B$33 "
            f"+ OFFSET({col}17, 0, -1) * (1 - $B$16 * {S_CHURN})"
        ),
        fmt=NUM)
    s.label(18, "Average headset subscriptions (year)", units="subs",
            source="(start + end) / 2; drives subscription revenue")
    s.year_formulas(18,
        lambda col, i: f"=({col}17 + OFFSET({col}17, 0, -1)) / 2", fmt=NUM)
    s.label(19, "Cumulative headsets sold (end of year)", units="headsets")
    s.year_formulas(19,
        lambda col, i: f"={delayed_units(col)} * {S_ADOPTION} + OFFSET({col}19, 0, -1)",
        fmt=NUM)

    # ---- Section 2: Hardware pricing ----
    s.section(21, "2. HARDWARE PRICING")
    s.label(22, "Hardware ASP", units="£/unit",
            source="ch.9 §9.4: mid of £250–350 range; allow time-varying")
    s.years(22, [300] * N_YEARS, fmt=GBP)
    s.label(23, "Returns allowance", units="% of hardware revenue",
            source="ch.9 §Warranty and Returns: DOA + trial-period returns only; in-warranty repairs covered separately by warranty provision")
    s.scalar(23, 0.02, fmt=PCT)

    # ---- Section 3: GTM channel mix, subscription & marketing split ----
    s.section(24, "3. GTM CHANNEL MIX, SUBSCRIPTION & MARKETING SPLIT")
    s.label(26, "Headsets per institutional buyer", units="headsets/buyer",
            source="ch.10 §10.2: schools/clubs each cover multiple sports")
    s.scalar(26, 3, fmt=INT)
    s.label(27, "Subscription price per headset", units="£/headset/yr",
            source="ch.10 §10.2: flat £600/yr; per-headset billing")
    s.scalar(27, 600, fmt=GBP)
    s.label(28, "Institutional buyers (schools + clubs)", units="buyers",
            source="ch.10 §10.2 (Phase 2 paid + Phase 3 institutional, scaling through Y7)")
    s.years(28, [0, 0, 30, 90, 150, 400, 750], fmt=NUM, prov=True)
    s.label(29, "DTC buyers (direct-to-consumer, no subscription)", units="buyers",
            source="ch.10 §10.3 (3% / 4% / 6% conversion ramp on effective exposure)")
    s.years(29, [0, 0, 0, 0, 1041, 1954, 5325], fmt=NUM, prov=True)
    s.label(30, "Institutional channel marketing", units="£/yr",
            source="B2B sales floor: conferences, sales reps, materials")
    s.years(30, [5000, 10000, 25000, 50000, 50000, 75000, 100000], fmt=GBP, prov=True)
    s.label(31, "DTC channel marketing", units="£/yr",
            source="Parent acquisition: digital, social, content marketing")
    s.years(31, [0, 0, 0, 0, 100000, 225000, 400000], fmt=GBP, prov=True)

    s.label(32, "Subscription ARPU per headset", units="£/sub/yr", total=True,
            source="= $B$27 (single-tier per-headset pricing)")
    s.year_formulas(32, lambda col, i: f"=$B$27", fmt=GBP, total=True)
    s.label(33, "Subscription attach (per institutional headset)",
            units="% of inst. headsets",
            source="100% by construction; DTC excluded in row 17 by using delayed_inst")
    s.scalar(33, 1.0, fmt=PCT)

    # ---- Section 4: Other revenue lines ----
    s.section(34, "4. OTHER REVENUE LINES   (toggle on/off)")
    s.label(35, "Research licences   (1 = on, 0 = off)")
    s.scalar(35, 0, fmt=INT)
    s.label(36, "Research licence sites per year", units="sites")
    s.years(36, [2, 5, 8, 10, 12, 14, 15], fmt=NUM)
    s.label(37, "Research licence £ per site per year", units="£/site/yr")
    s.scalar(37, 3000, fmt=GBP)

    s.label(38, "Accessories & consumables   (1 = on, 0 = off)")
    s.scalar(38, 0, fmt=INT)
    s.label(39, "Accessories £/yr per active subscriber", units="£/sub/yr",
            source="Wipes, replacement facial interfaces, straps")
    s.scalar(39, 20, fmt=GBP)
    s.label(40, "Accessories gross margin %", units="%",
            source="Markup on consumables; assumed similar to medical disposables")
    s.scalar(40, 0.60, fmt=PCT)

    s.label(41, "Extended warranty   (1 = on, 0 = off)")
    s.scalar(41, 0, fmt=INT)
    s.label(42, "Extended warranty attach % at point of sale", units="%")
    s.scalar(42, 0.15, fmt=PCT)
    s.label(43, "Extended warranty £ per attached unit", units="£/unit")
    s.scalar(43, 45, fmt=GBP)
    s.label(44, "Extended warranty gross margin %", units="%",
            source="ch.9 §9.5: 40–50% range")
    s.scalar(44, 0.45, fmt=PCT)

    # ---- Section 5: COGS ----
    s.section(45, "5. COST OF GOODS SOLD")
    s.label(46, "Unit cost — P1 (~10 units)", units="£/unit",
            source="ch.9 tab:unit-cost-stack P1 midpoint (305-380)")
    s.scalar(46, 340, fmt=GBP)
    s.label(47, "Unit cost — P2 (~500 units)", units="£/unit",
            source="ch.9 tab:unit-cost-stack P2 midpoint (252-318)")
    s.scalar(47, 285, fmt=GBP)
    s.label(48, "Unit cost — P3 (~10k units, RK3588S migration)", units="£/unit",
            source="ch.9 tab:unit-cost-stack P3 midpoint (138-209)")
    s.scalar(48, 174, fmt=GBP)
    s.label(49, "Effective unit cost per year", units="£/unit",
            source="CHOOSE(phase, P1, P2, P3) × BOM scenario multiplier")
    s.year_formulas(49,
        lambda col, i: f"=CHOOSE({col}7, $B$46, $B$47, $B$48) * {S_BOM}", fmt=GBP)
    s.label(50, "Subscription COGS as % of subscription revenue", units="%",
            source="hosting, support tooling, data egress")
    s.scalar(50, 0.15, fmt=PCT)

    # ---- Section 6: Headcount & payroll ----
    s.section(52, "6. HEADCOUNT & PAYROLL")
    s.cell(53, 1, "Role", font=font_subheader)
    s.cell(53, 2, "Salary £/yr", font=font_subheader,
           align=Alignment(horizontal="center"))
    for i, y in enumerate(YEAR_LABELS):
        s.cell(53, 3 + i, "FTE " + y, font=font_subheader,
               align=Alignment(horizontal="center"))

    roles = [
        ("CEO / commercial lead",            50000, [1,   1,   1, 1,   1,   1,   1]),
        ("Engineering (HW / firmware / CV)", 55000, [3,   3,   4, 5,   6,   7,   8]),
        ("Regulatory & quality",             65000, [0,   0.5, 1, 1.5, 2,   2,   2]),
        ("Clinical research",                40000, [0,   0.5, 1, 1,   1,   1,   1]),
        ("Sales / business development",     48000, [0,   0,   1, 2,   3,   3,   3]),
        ("Operations / supply chain",        45000, [0,   0,   0, 1,   1,   1,   1]),
        ("Customer support",                 32000, [0,   0,   0, 0.5, 1,   1.5, 2]),
        ("(spare slot)",                     0,     [0]*N_YEARS),
    ]
    for i, (name, sal, fte) in enumerate(roles):
        r = 54 + i
        s.cell(r, 1, name, font=font_input)
        s.cell(r, 2, sal, font=font_input, fmt=GBP)
        for j, n in enumerate(fte):
            s.cell(r, 3 + j, n, font=font_input, fmt=FTE)

    s.label(62, "Total FTE", subheader=True)
    s.year_formulas(62, lambda col, i: f"=SUM({col}54:{col}61)", fmt=FTE)
    s.label(63, "Base payroll  (cash salaries)", units="£")
    s.year_formulas(63,
        lambda col, i: f"=SUMPRODUCT($B$54:$B$61, {col}54:{col}61)", fmt=GBP)
    s.label(64, "Founder cash adjustment", units="£",
            source="Negative = founders take partial cash; affects P1–P2 only")
    s.years(64, [-30000, -30000, 0, 0, 0, 0, 0], fmt=GBP)
    s.label(65, "On-cost loading %", units="%",
            source="ch.9 §9.5: NI 12–14% + pension 3% + benefits ~2%")
    s.scalar(65, 0.19, fmt=PCT)
    s.label(66, "Loaded payroll  (incl. on-cost)", units="£",
            source="(base + founder adj) × (1 + on-cost) × payroll scenario multiplier")
    s.year_formulas(66,
        lambda col, i: f"=({col}63 + {col}64) * (1 + $B$65) * {S_PAYROLL}", fmt=GBP)
    s.label(67, "Sales commission %", units="% of new ARR",
            source="Applied to gross new subscription revenue")
    s.scalar(67, 0.15, fmt=PCT)
    s.label(68, "Sales commission £", units="£",
            source="institutional buyers × headsets/inst × adoption × attach × ARPU × commission %")
    s.year_formulas(68,
        lambda col, i: (
            f"={delayed_inst(col)} * $B$26 * {S_ADOPTION} * $B$33 * {col}32 * $B$67"
        ),
        fmt=GBP)
    s.label(69, "Total people cost", units="£", total=True,
            source="= loaded payroll + sales commission")
    s.year_formulas(69, lambda col, i: f"={col}66 + {col}68", fmt=GBP, total=True)

    # ---- Section 7: Other opex ----
    s.section(71, "7. OTHER OPERATING EXPENSES")
    s.label(72, "R&D non-payroll", units="£",
            source="Prototyping components, dev tools, CAD/sim licences")
    s.years(72, [15000, 20000, 25000, 30000, 30000, 30000, 30000], fmt=GBP)
    s.label(73, "Marketing (derived)", units="£",
            source="= institutional channel marketing + DTC channel marketing (rows 30 + 31)")
    s.year_formulas(73, lambda col, i: f"={col}30 + {col}31", fmt=GBP)
    s.label(74, "Travel & subsistence", units="£")
    s.years(74, [8000, 12000, 25000, 40000, 55000, 70000, 85000], fmt=GBP)
    s.label(75, "Professional fees", units="£",
            source="Legal, accounting, IP, audit")
    s.years(75, [10000, 15000, 30000, 40000, 50000, 60000, 70000], fmt=GBP)
    s.label(76, "IT and software, cloud infrastructure", units="£",
            source="SaaS licences, dev infrastructure, cloud hosting (scales with installed base)")
    s.years(76, [6000, 10000, 20000, 35000, 55000, 75000, 95000], fmt=GBP)
    s.label(77, "Total other opex", units="£", total=True)
    s.year_formulas(77, lambda col, i: f"=SUM({col}72:{col}76)", fmt=GBP, total=True)

    # ---- Section 8: Facilities ----
    s.section(79, "8. FACILITIES")
    s.label(80, "Facilities annual cost", units="£",
            source="ch.9 §9.5: free incubator → coworking → office")
    s.years(80, [0, 20000, 24000, 35000, 45000, 55000, 65000], fmt=GBP)

    # ---- Section 9: Insurance ----
    s.section(82, "9. INSURANCE")
    s.label(83, "Insurance annual cost", units="£",
            source="ch.9 tab:ent-ops-insurance — phased programme")
    s.years(83, [6000, 12000, 20000, 28000, 36000, 44000, 50000], fmt=GBP)

    # ---- Section 10: Regulatory & clinical ----
    s.section(85, "10. REGULATORY & CLINICAL   [PROVISIONAL → ch.7]")
    s.label(86, "One-off regulatory & clinical (UKCA route)", units="£",
            source="QMS + biocompat + EMC + NB fees + clinical investigation")
    s.years(86, [0, 130000, 140000, 0, 0, 0, 0], fmt=GBP, prov=True)
    s.label(87, "Ongoing post-market & NB surveillance", units="£",
            source="PMS, periodic safety updates, NB annual surveillance")
    s.years(87, [0, 0, 0, 55000, 55000, 60000, 65000], fmt=GBP, prov=True)
    s.label(88, "Total regulatory & clinical", units="£", total=True)
    s.year_formulas(88, lambda col, i: f"={col}86 + {col}87", fmt=GBP, total=True)

    # ---- Section 11: Capex & depreciation ----
    s.section(90, "11. CAPEX & DEPRECIATION")
    s.cell(91, 1, "Capex item", font=font_subheader)
    s.cell(91, 2, "Amount £",  font=font_subheader, align=Alignment(horizontal="center"))
    s.cell(91, 3, "Period",    font=font_subheader, align=Alignment(horizontal="center"))
    s.cell(91, 4, "Life (yrs)", font=font_subheader, align=Alignment(horizontal="center"))

    capex = [
        ("Bench equipment, ESD, basic instruments",    4000,  1, 5),
        ("Desktop 3D printer + accessories",           2500,  1, 5),
        ("Optical alignment fixtures (in-house)",      6000,  2, 5),
        ("Test fixtures (formal, jig-assisted)",       15000, 3, 5),
        ("Office fit-out (lab area, desks, security)", 20000, 4, 5),
        ("Aluminium injection moulding tool",          12000, 5, 5),
        ("Calibration jigs for contract manufacturer", 18000, 5, 5),
        ("End-of-line test rig upgrade",               12000, 6, 5),
        ("Steel injection moulding tool",              22000, 7, 7),
        ("(spare slot)",                               0,     1, 5),
    ]
    for i, (item, amt, per, life) in enumerate(capex):
        r = 92 + i
        s.cell(r, 1, item, font=font_input)
        s.cell(r, 2, amt,  font=font_input, fmt=GBP)
        s.cell(r, 3, per,  font=font_input, fmt=INT, align=Alignment(horizontal="center"))
        s.cell(r, 4, life, font=font_input, fmt=INT, align=Alignment(horizontal="center"))

    s.label(103, "Capex per year", units="£", subheader=True)
    s.year_formulas(103,
        lambda col, i: f"=SUMIFS($B$92:$B$101, $C$92:$C$101, {col}5)", fmt=GBP)
    s.label(104, "Cumulative capex", units="£")
    s.year_formulas(104,
        lambda col, i: f"={col}103 + OFFSET({col}104, 0, -1)", fmt=GBP)
    s.label(105, "Depreciation per year", units="£",
            source="Straight-line over each item's life, from period of acquisition")
    s.year_formulas(105,
        lambda col, i: (
            f"=SUMPRODUCT(($C$92:$C$101 <= {col}5) * "
            f"({col}5 < $C$92:$C$101 + $D$92:$D$101) * "
            f"($B$92:$B$101 / $D$92:$D$101))"
        ), fmt=GBP)
    s.label(106, "Accumulated depreciation", units="£")
    s.year_formulas(106,
        lambda col, i: f"={col}105 + OFFSET({col}106, 0, -1)", fmt=GBP)
    s.label(107, "Net PP&E", units="£", total=True,
            source="= cumulative capex − accumulated depreciation")
    s.year_formulas(107, lambda col, i: f"={col}104 - {col}106", fmt=GBP, total=True)

    # ---- Section 12: Working capital ----
    s.section(109, "12. WORKING CAPITAL ASSUMPTIONS")
    s.label(110, "Inventory days  (DIO, on COGS)", units="days")
    s.scalar(110, 75, fmt=INT)
    s.label(111, "Receivable days  (DSO, on revenue)", units="days")
    s.scalar(111, 45, fmt=INT)
    s.label(112, "Payable days  (DPO, on COGS + opex ex-payroll)", units="days")
    s.scalar(112, 45, fmt=INT)
    s.label(113, "Warranty provision % of hardware revenue", units="%",
            source="ch.9 §9.5: 2–5% range")
    s.scalar(113, 0.03, fmt=PCT)

    # ---- Section 13: Tax ----
    s.section(115, "13. TAX")
    s.label(116, "UK Corporation Tax rate", units="%",
            source="25% main rate from April 2023")
    s.scalar(116, 0.25, fmt=PCT)
    s.label(117, "R&D tax credit   (1 = on, 0 = off)")
    s.scalar(117, 1, fmt=INT)
    s.label(118, "R&D credit effective rate", units="% of qualifying R&D",
            source="UK SME scheme post April 2024 (~18.6% effective benefit)")
    s.scalar(118, 0.186, fmt=PCT)
    s.label(119, "Qualifying R&D — % of engineering payroll", units="%")
    s.scalar(119, 0.60, fmt=PCT)
    s.label(120, "Qualifying R&D — % of regulatory payroll", units="%")
    s.scalar(120, 0.40, fmt=PCT)
    s.label(121, "Qualifying R&D — % of clinical payroll", units="%")
    s.scalar(121, 0.30, fmt=PCT)
    s.label(122, "Qualifying R&D — % of clinical investigation costs", units="%")
    s.scalar(122, 0.50, fmt=PCT)
    s.label(123, "Qualifying R&D — % of R&D non-payroll spend", units="%")
    s.scalar(123, 1.00, fmt=PCT)
    s.label(124, "Interest on cash balance", units="% p.a.",
            source="Business savings rate; applied to opening cash in P&L")
    s.scalar(124, 0.03, fmt=PCT)

    # ---- Section 14: Equity injections ----
    s.section(125, "14. EQUITY INJECTIONS   [PLUG]")
    s.label(126, "Equity injections", units="£",
            source="Mechanical plug — sized to keep closing cash ≥ buffer at all times")
    s.years(126, [900000, 0, 1500000, 0, 1500000, 0, 0], fmt=GBP)
    s.label(127, "Minimum cash buffer", units="£")
    s.scalar(127, 100000, fmt=GBP)

    # Freeze panes so the year header rows stay visible while scrolling
    ws.freeze_panes = "C8"


# ---------------------------------------------------------------------------
# UNIT ECONOMICS TAB
# ---------------------------------------------------------------------------
def build_unit_economics(wb):
    ws = wb.create_sheet("Unit economics")
    s = Sheet(ws)
    A = "Assumptions!"

    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 14
    for col in YEAR_COLS:
        ws.column_dimensions[col].width = 13
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 42

    # ---- Title ----
    s.cell(1, 1, "Unit Economics — Per-customer profitability", font=font_title)
    ws.merge_cells("A1:M1")
    s.cell(2, 1,
           "All values driven from the Assumptions tab. Y6–Y7 are the steady-state reference (P3 mature, post-breakeven).",
           font=font_note)
    ws.merge_cells("A2:M2")

    # ---- Year header rows (linked from Assumptions) ----
    s.label(4, "Year")
    for i, col in enumerate(YEAR_COLS):
        c = s.cell(4, 3 + i, f"={A}{col}4",
                   font=font_year_label, fill=fill_year,
                   align=Alignment(horizontal="center"))
    s.label(5, "Period number")
    s.year_formulas(5, lambda col, i: f"={A}{col}5", fmt=INT)
    s.label(6, "Phase")
    s.year_formulas(6, lambda col, i: f"={A}{col}7", fmt=INT)

    # ---- Section 1: Hardware economics ----
    s.section(8, "1. HARDWARE ECONOMICS  (per unit)")
    s.label(9, "Hardware ASP", units="£/unit")
    for i, col in enumerate(YEAR_COLS):
        s.cell(9, 3 + i, f"={A}{col}22", font=font_link, fmt=GBP)
    s.label(10, "Effective unit cost", units="£/unit")
    for i, col in enumerate(YEAR_COLS):
        s.cell(10, 3 + i, f"={A}{col}49", font=font_link, fmt=GBP)
    s.label(11, "Hardware contribution per unit", units="£/unit", total=True)
    s.year_formulas(11, lambda col, i: f"={col}9 - {col}10", fmt=GBP, total=True)
    s.label(12, "Hardware gross margin %", units="%")
    s.year_formulas(12, lambda col, i: f"=IFERROR({col}11/{col}9, 0)", fmt=PCT)

    # ---- Section 2: Subscription economics ----
    s.section(14, "2. SUBSCRIPTION ECONOMICS  (per active customer per year)")
    s.label(15, "Blended ARPU", units="£/sub/yr")
    for i, col in enumerate(YEAR_COLS):
        s.cell(15, 3 + i, f"={A}{col}32", font=font_link, fmt=GBP)
    s.label(16, "Subscription COGS rate", units="%")
    for i, col in enumerate(YEAR_COLS):
        s.cell(16, 3 + i, f"={A}$B$50", font=font_link, fmt=PCT)
    s.label(17, "Subscription contribution", units="£/sub/yr", total=True)
    s.year_formulas(17, lambda col, i: f"={col}15 * (1 - {col}16)", fmt=GBP, total=True)
    s.label(18, "Subscription gross margin %", units="%")
    s.year_formulas(18, lambda col, i: f"=IFERROR({col}17/{col}15, 0)", fmt=PCT)

    # ---- Section 3: Accessories ----
    s.section(20, "3. ACCESSORIES & CONSUMABLES  (per active customer per year)")
    s.label(21, "Accessories enabled  (1 = on, 0 = off)")
    for i, col in enumerate(YEAR_COLS):
        s.cell(21, 3 + i, f"={A}$B$38", font=font_link, fmt=INT)
    s.label(22, "Accessories revenue per active sub", units="£/sub/yr")
    for i, col in enumerate(YEAR_COLS):
        s.cell(22, 3 + i, f"={col}21 * {A}$B$39", font=font_formula, fmt=GBP)
    s.label(23, "Accessories gross margin %", units="%")
    for i, col in enumerate(YEAR_COLS):
        s.cell(23, 3 + i, f"={A}$B$40", font=font_link, fmt=PCT)
    s.label(24, "Accessories contribution per active sub", units="£/sub/yr", total=True)
    s.year_formulas(24, lambda col, i: f"={col}22 * {col}23", fmt=GBP, total=True)

    # ---- Section 4: Extended warranty ----
    s.section(26, "4. EXTENDED WARRANTY  (one-off, per new customer at sale)")
    s.label(27, "Warranty enabled  (1 = on, 0 = off)")
    for i, col in enumerate(YEAR_COLS):
        s.cell(27, 3 + i, f"={A}$B$41", font=font_link, fmt=INT)
    s.label(28, "Attach % at sale", units="%")
    for i, col in enumerate(YEAR_COLS):
        s.cell(28, 3 + i, f"={A}$B$42", font=font_link, fmt=PCT)
    s.label(29, "Warranty £ per attached unit", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(29, 3 + i, f"={A}$B$43", font=font_link, fmt=GBP)
    s.label(30, "Warranty gross margin %", units="%")
    for i, col in enumerate(YEAR_COLS):
        s.cell(30, 3 + i, f"={A}$B$44", font=font_link, fmt=PCT)
    s.label(31, "Warranty contribution per new unit", units="£/unit", total=True)
    s.year_formulas(31,
        lambda col, i: f"={col}27 * {col}28 * {col}29 * {col}30", fmt=GBP, total=True)

    # ---- Section 5: Total annual economics per active customer ----
    s.section(33, "5. TOTAL ANNUAL ECONOMICS  (per active customer)")
    s.label(34, "Total annual revenue per active customer", units="£/sub/yr",
            source="= ARPU + accessories revenue")
    s.year_formulas(34, lambda col, i: f"={col}15 + {col}22", fmt=GBP)
    s.label(35, "Total annual contribution per active customer", units="£/sub/yr", total=True,
            source="= sub contribution + accessories contribution")
    s.year_formulas(35, lambda col, i: f"={col}17 + {col}24", fmt=GBP, total=True)

    # ---- Section 6: LTV ----
    s.section(37, "6. CUSTOMER LIFETIME VALUE (LTV)")
    s.label(38, "Annual subscription churn", units="%")
    for i, col in enumerate(YEAR_COLS):
        s.cell(38, 3 + i, f"={A}$B$16", font=font_link, fmt=PCT)
    s.label(39, "Implied avg customer lifetime", units="years",
            source="= 1 / churn (perpetuity equivalent)")
    s.year_formulas(39, lambda col, i: f"=IFERROR(1/{col}38, 0)", fmt='0.0;(0.0);"-"')
    s.label(40, "3-year LTV factor (sub side)", units="multiplier",
            source="= 1 + (1−c) + (1−c)²")
    s.year_formulas(40,
        lambda col, i: f"=1 + (1-{col}38) + (1-{col}38)^2",
        fmt='0.00;(0.00);"-"')
    s.label(41, "3-year sub + accessories LTV", units="£/customer")
    s.year_formulas(41, lambda col, i: f"={col}35 * {col}40", fmt=GBP)
    s.label(42, "3-year total LTV per customer", units="£/customer", total=True,
            source="= HW contribution + warranty contribution + 3-yr sub+acc LTV")
    s.year_formulas(42, lambda col, i: f"={col}11 + {col}31 + {col}41", fmt=GBP, total=True)
    s.label(43, "Steady-state sub + accessories LTV", units="£/customer",
            source="= annual contribution / churn (perpetuity)")
    s.year_formulas(43, lambda col, i: f"=IFERROR({col}35/{col}38, 0)", fmt=GBP)
    s.label(44, "Steady-state total LTV per customer", units="£/customer", total=True)
    s.year_formulas(44, lambda col, i: f"={col}11 + {col}31 + {col}43", fmt=GBP, total=True)

    # ---- Section 7: CAC by channel ----
    s.section(46, "7. CUSTOMER ACQUISITION COST  (by channel)")
    s.label(47, "Institutional channel marketing", units="£/yr")
    for i, col in enumerate(YEAR_COLS):
        s.cell(47, 3 + i, f"={A}{col}30", font=font_link, fmt=GBP)
    s.label(48, "Sales commission (institutional only)", units="£/yr")
    for i, col in enumerate(YEAR_COLS):
        s.cell(48, 3 + i, f"={A}{col}68", font=font_link, fmt=GBP)
    s.label(49, "Institutional acquisition spend", units="£/yr")
    s.year_formulas(49, lambda col, i: f"={col}47 + {col}48", fmt=GBP)
    s.label(50, "Institutional buyers", units="buyers")
    for i, col in enumerate(YEAR_COLS):
        s.cell(50, 3 + i, f"={A}{col}28", font=font_link, fmt=NUM)
    s.label(51, "CAC per institutional customer", units="£/customer", total=True,
            source="= institutional acquisition spend / institutional buyers")
    s.year_formulas(51, lambda col, i: f"=IFERROR({col}49/{col}50, 0)", fmt=GBP, total=True)

    s.label(52, "DTC channel marketing", units="£/yr")
    for i, col in enumerate(YEAR_COLS):
        s.cell(52, 3 + i, f"={A}{col}31", font=font_link, fmt=GBP)
    s.label(53, "DTC buyers", units="buyers")
    for i, col in enumerate(YEAR_COLS):
        s.cell(53, 3 + i, f"={A}{col}29", font=font_link, fmt=NUM)
    s.label(54, "CAC per DTC customer", units="£/customer", total=True,
            source="= DTC marketing / DTC buyers (no commission on DTC)")
    s.year_formulas(54, lambda col, i: f"=IFERROR({col}52/{col}53, 0)", fmt=GBP, total=True)

    # ---- Section 8: LTV per customer & LTV:CAC by channel ----
    s.section(56, "8. LTV PER CUSTOMER & LTV : CAC  (by channel)")
    s.label(57, "3-year LTV per institutional customer", units="£/customer", total=True,
            source="= 3 headsets × (HW contribution + 3-yr sub LTV per headset)")
    s.year_formulas(57,
        lambda col, i: f"={A}$B$26 * ({col}11 + {col}41)",
        fmt=GBP, total=True)
    s.label(58, "LTV per DTC customer (one-off)", units="£/customer", total=True,
            source="= hardware contribution per headset (no subscription)")
    s.year_formulas(58, lambda col, i: f"={col}11", fmt=GBP, total=True)
    s.label(59, "LTV : CAC ratio  (institutional, 3-year)", units="ratio", total=True,
            source="Industry benchmark > 3.0x")
    s.year_formulas(59,
        lambda col, i: f"=IFERROR({col}57/{col}51, 0)",
        fmt='0.0"x";(0.0"x");"-"', total=True)
    s.label(60, "LTV : CAC ratio  (DTC)", units="ratio", total=True,
            source="DTC must clear 1.0x to be self-funding at acquisition")
    s.year_formulas(60,
        lambda col, i: f"=IFERROR({col}58/{col}54, 0)",
        fmt='0.0"x";(0.0"x");"-"', total=True)
    s.label(61, "Payback period (institutional)", units="months",
            source="= CAC × 12 / annual contribution per institutional customer (3 headsets)")
    s.year_formulas(61,
        lambda col, i: f"=IFERROR({col}51*12/({A}$B$26 * {col}35), 0)",
        fmt='0.0;(0.0);"-"')

    # ---- Section 9: Steady-state summary panel (Y7 reference) ----
    s.section(63, "9. STEADY-STATE SUMMARY  (Y7 reference)")
    panel = [
        ("Hardware contribution per headset",             "K11", GBP),
        ("Subscription contribution per headset/yr",      "K17", GBP),
        ("3-year LTV per institutional customer",         "K57", GBP),
        ("LTV per DTC customer (hardware only)",          "K58", GBP),
        ("CAC per institutional customer",                "K51", GBP),
        ("CAC per DTC customer",                          "K54", GBP),
        ("LTV : CAC  (institutional, 3-year)",            "K59", '0.0"x";(0.0"x");"-"'),
        ("LTV : CAC  (DTC)",                              "K60", '0.0"x";(0.0"x");"-"'),
        ("Payback period (institutional, months)",        "K61", '0.0;(0.0);"-"'),
    ]
    for i, (lbl, ref, fmt) in enumerate(panel):
        r = 64 + i
        s.cell(r, 1, lbl, font=font_subheader, fill=fill_subsection)
        s.cell(r, 2, f"={ref}", font=font_link, fmt=fmt, fill=fill_subsection)
        for col in range(3, LAST_COL + 1):
            s.ws.cell(r, col).fill = fill_subsection

    ws.freeze_panes = "C8"


# ---------------------------------------------------------------------------
# P&L TAB
# ---------------------------------------------------------------------------
def build_pnl(wb):
    ws = wb.create_sheet("P&L")
    s = Sheet(ws)
    A = "Assumptions!"

    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 13
    for col in YEAR_COLS:
        ws.column_dimensions[col].width = 13
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 42

    # ---- Title ----
    s.cell(1, 1, "Profit & Loss — 7-year projection", font=font_title)
    ws.merge_cells("A1:M1")
    s.cell(2, 1,
           "Driven from the Assumptions tab. Read the EBITDA line for the headline shape; "
           "read R&D credit + corp tax for the cash impact of loss-making years.",
           font=font_note)
    ws.merge_cells("A2:M2")

    # ---- Year header rows (linked) ----
    s.label(4, "Year")
    for i, col in enumerate(YEAR_COLS):
        s.cell(4, 3 + i, f"={A}{col}4",
               font=font_year_label, fill=fill_year,
               align=Alignment(horizontal="center"))
    s.label(5, "Period number")
    s.year_formulas(5, lambda col, i: f"={A}{col}5", fmt=INT)
    s.label(6, "Phase")
    s.year_formulas(6, lambda col, i: f"={A}{col}7", fmt=INT)

    # ---- Section 1: Revenue ----
    s.section(8, "1. REVENUE")
    s.label(9, "Hardware revenue", units="£",
            source="delay-shifted new units × adoption scenario × ASP × (1 − returns allowance)")
    s.year_formulas(9,
        lambda col, i: f"={delayed_units(col, A)} * {S_ADOPTION} * {A}{col}22 * (1 - {A}$B$23)",
        fmt=GBP)
    s.label(10, "Subscription revenue", units="£",
            source="average subscribers × blended ARPU")
    s.year_formulas(10, lambda col, i: f"={A}{col}18 * {A}{col}32", fmt=GBP)
    s.label(11, "Research licence revenue", units="£",
            source="toggle × sites × £/site")
    s.year_formulas(11,
        lambda col, i: f"={A}$B$35 * {A}{col}36 * {A}$B$37", fmt=GBP)
    s.label(12, "Accessories & consumables revenue", units="£",
            source="toggle × avg subs × £/sub/yr")
    s.year_formulas(12,
        lambda col, i: f"={A}$B$38 * {A}{col}18 * {A}$B$39", fmt=GBP)
    s.label(13, "Extended warranty revenue", units="£",
            source="toggle × delay-shifted new units × adoption × attach × £/unit")
    s.year_formulas(13,
        lambda col, i: f"={A}$B$41 * {delayed_units(col, A)} * {S_ADOPTION} * {A}$B$42 * {A}$B$43",
        fmt=GBP)
    s.label(14, "TOTAL REVENUE", units="£", total=True)
    s.year_formulas(14, lambda col, i: f"=SUM({col}9:{col}13)", fmt=GBP, total=True)

    # ---- Section 2: COGS ----
    s.section(16, "2. COST OF GOODS SOLD")
    s.label(17, "Hardware COGS", units="£",
            source="delay-shifted new units × adoption × effective unit cost (BOM applied in row 49)")
    s.year_formulas(17,
        lambda col, i: f"={delayed_units(col, A)} * {S_ADOPTION} * {A}{col}49",
        fmt=GBP)
    s.label(18, "Subscription COGS", units="£",
            source="hosting / support tooling — sub rev × COGS rate")
    s.year_formulas(18, lambda col, i: f"={col}10 * {A}$B$50", fmt=GBP)
    s.label(19, "Accessories COGS", units="£")
    s.year_formulas(19, lambda col, i: f"={col}12 * (1 - {A}$B$40)", fmt=GBP)
    s.label(20, "Extended warranty COGS", units="£")
    s.year_formulas(20, lambda col, i: f"={col}13 * (1 - {A}$B$44)", fmt=GBP)
    s.label(21, "TOTAL COGS", units="£", total=True)
    s.year_formulas(21, lambda col, i: f"=SUM({col}17:{col}20)", fmt=GBP, total=True)

    s.label(23, "GROSS PROFIT", units="£", total=True)
    s.year_formulas(23, lambda col, i: f"={col}14 - {col}21", fmt=GBP, total=True)
    s.label(24, "Gross margin %", units="%")
    s.year_formulas(24, lambda col, i: f"=IFERROR({col}23/{col}14, 0)", fmt=PCT)

    # ---- Helper: payroll allocation formula ----
    def alloc(salary_ref, fte_col_ref):
        # role base ÷ total base × total loaded payroll
        return (
            f"=IFERROR({A}{salary_ref} * {A}{fte_col_ref} / {A}{fte_col_ref[0]}63, 0)"
            f" * {A}{fte_col_ref[0]}66"
        )

    # ---- Section 3: OpEx — R&D ----
    s.section(26, "3. OPERATING EXPENSES — R&D")
    s.label(27, "Engineering payroll (allocated)", units="£",
            source="(Eng base ÷ total base) × loaded payroll")
    s.year_formulas(27, lambda col, i: alloc("$B$55", f"{col}55"), fmt=GBP)
    s.label(28, "R&D non-payroll", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(28, 3 + i, f"={A}{col}72", font=font_link, fmt=GBP)
    s.label(29, "TOTAL R&D", units="£", total=True)
    s.year_formulas(29, lambda col, i: f"={col}27 + {col}28", fmt=GBP, total=True)

    # ---- Section 4: OpEx — Regulatory & Clinical ----
    s.section(31, "4. OPERATING EXPENSES — REGULATORY & CLINICAL")
    s.label(32, "Regulatory & quality payroll (allocated)", units="£")
    s.year_formulas(32, lambda col, i: alloc("$B$56", f"{col}56"), fmt=GBP)
    s.label(33, "Clinical research payroll (allocated)", units="£")
    s.year_formulas(33, lambda col, i: alloc("$B$57", f"{col}57"), fmt=GBP)
    s.label(34, "One-off regulatory & clinical", units="£",
            source="QMS, NB fees, biocompat, EMC, clinical investigation × reg scenario")
    for i, col in enumerate(YEAR_COLS):
        s.cell(34, 3 + i,
               f"={A}{col}86 * {S_REG}",
               font=font_formula, fmt=GBP)
    s.label(35, "Ongoing PMS & NB surveillance", units="£",
            source="Post-market obligations × reg scenario")
    for i, col in enumerate(YEAR_COLS):
        s.cell(35, 3 + i,
               f"={A}{col}87 * {S_REG}",
               font=font_formula, fmt=GBP)
    s.label(36, "TOTAL REGULATORY & CLINICAL", units="£", total=True)
    s.year_formulas(36, lambda col, i: f"=SUM({col}32:{col}35)", fmt=GBP, total=True)

    # ---- Section 5: OpEx — Sales & Marketing ----
    s.section(38, "5. OPERATING EXPENSES — SALES & MARKETING")
    s.label(39, "Sales / BD payroll (allocated)", units="£")
    s.year_formulas(39, lambda col, i: alloc("$B$58", f"{col}58"), fmt=GBP)
    s.label(40, "Customer support payroll (allocated)", units="£")
    s.year_formulas(40, lambda col, i: alloc("$B$60", f"{col}60"), fmt=GBP)
    s.label(41, "Sales commission", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(41, 3 + i, f"={A}{col}68", font=font_link, fmt=GBP)
    s.label(42, "Marketing", units="£",
            source="Marketing spend × marketing scenario")
    for i, col in enumerate(YEAR_COLS):
        s.cell(42, 3 + i,
               f"={A}{col}73 * {S_MARKETING}",
               font=font_formula, fmt=GBP)
    s.label(43, "TOTAL SALES & MARKETING", units="£", total=True)
    s.year_formulas(43, lambda col, i: f"=SUM({col}39:{col}42)", fmt=GBP, total=True)

    # ---- Section 6: OpEx — G&A ----
    s.section(45, "6. OPERATING EXPENSES — GENERAL & ADMINISTRATIVE")
    s.label(46, "CEO / commercial payroll (allocated)", units="£")
    s.year_formulas(46, lambda col, i: alloc("$B$54", f"{col}54"), fmt=GBP)
    s.label(47, "Operations / supply chain payroll (allocated)", units="£")
    s.year_formulas(47, lambda col, i: alloc("$B$59", f"{col}59"), fmt=GBP)
    s.label(48, "Spare role payroll (allocated)", units="£")
    s.year_formulas(48, lambda col, i: alloc("$B$61", f"{col}61"), fmt=GBP)
    s.label(49, "Travel & subsistence", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(49, 3 + i, f"={A}{col}74", font=font_link, fmt=GBP)
    s.label(50, "Professional fees", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(50, 3 + i, f"={A}{col}75", font=font_link, fmt=GBP)
    s.label(51, "IT and software", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(51, 3 + i, f"={A}{col}76", font=font_link, fmt=GBP)
    s.label(52, "Facilities", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(52, 3 + i, f"={A}{col}80", font=font_link, fmt=GBP)
    s.label(53, "Insurance", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(53, 3 + i, f"={A}{col}83", font=font_link, fmt=GBP)
    s.label(54, "TOTAL G&A", units="£", total=True)
    s.year_formulas(54, lambda col, i: f"=SUM({col}46:{col}53)", fmt=GBP, total=True)

    s.label(56, "TOTAL OPERATING EXPENSES", units="£", total=True)
    s.year_formulas(56,
        lambda col, i: f"={col}29 + {col}36 + {col}43 + {col}54",
        fmt=GBP, total=True)

    # ---- EBITDA / EBIT ----
    s.label(58, "EBITDA", units="£", total=True)
    s.year_formulas(58, lambda col, i: f"={col}23 - {col}56", fmt=GBP, total=True)
    s.label(59, "EBITDA margin %", units="%")
    s.year_formulas(59, lambda col, i: f"=IFERROR({col}58/{col}14, 0)", fmt=PCT)

    s.label(61, "Depreciation & amortisation", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(61, 3 + i, f"={A}{col}105", font=font_link, fmt=GBP)

    s.label(63, "EBIT (Operating profit / loss)", units="£", total=True)
    s.year_formulas(63, lambda col, i: f"={col}58 - {col}61", fmt=GBP, total=True)
    s.label(64, "+ Interest income on opening cash", units="£",
            source="Cash flow opening cash × Assumptions!$B$124")
    s.year_formulas(64,
        lambda col, i: f"='Cash flow'!{col}38 * {A}$B$124", fmt=GBP)

    # ---- Section 7: Tax ----
    s.section(65, "7. TAX  (with loss carry-forward and R&D credit)")
    s.label(66, "Cumulative losses brought forward", units="£",
            source="Carries unused trading losses indefinitely")
    s.year_formulas(66, lambda col, i: f"=OFFSET({col}66, 7, -1)", fmt=GBP)
    s.label(67, "Losses utilised this year", units="£",
            source="MIN(losses BF, max(0, EBIT + interest))")
    s.year_formulas(67,
        lambda col, i: f"=MIN({col}66, MAX(0, {col}63 + {col}64))", fmt=GBP)
    s.label(68, "Taxable profit", units="£",
            source="MAX(0, EBIT + interest − losses utilised)")
    s.year_formulas(68,
        lambda col, i: f"=MAX(0, {col}63 + {col}64 - {col}67)", fmt=GBP)
    s.label(69, "Corporation tax", units="£",
            source="taxable profit × 25%")
    s.year_formulas(69,
        lambda col, i: f"={col}68 * {A}$B$116", fmt=GBP)
    s.label(70, "Qualifying R&D spend", units="£",
            source="Eng×60% + Reg×40% + Clin×30% + R&D npp×100% + clin inv×50%")
    s.year_formulas(70,
        lambda col, i: (
            f"={col}27*{A}$B$119 + {col}32*{A}$B$120 + {col}33*{A}$B$121 + "
            f"{col}28*{A}$B$123 + {col}34*{A}$B$122"
        ), fmt=GBP)
    s.label(71, "R&D tax credit", units="£",
            source="−qualifying × effective rate × on/off; cash-back when loss-making")
    s.year_formulas(71,
        lambda col, i: f"=-{col}70 * {A}$B$118 * {A}$B$117", fmt=GBP)
    s.label(72, "Net tax expense / (credit)", units="£", total=True)
    s.year_formulas(72, lambda col, i: f"={col}69 + {col}71", fmt=GBP, total=True)
    s.label(73, "Cumulative losses carried forward", units="£",
            source="prior CF + this year's loss − losses utilised")
    s.year_formulas(73,
        lambda col, i: f"={col}66 + MAX(0, -({col}63 + {col}64)) - {col}67", fmt=GBP)

    # ---- Net income ----
    s.label(75, "NET INCOME", units="£", total=True)
    s.year_formulas(75, lambda col, i: f"={col}63 + {col}64 - {col}72", fmt=GBP, total=True)
    s.label(76, "Net margin %", units="%")
    s.year_formulas(76, lambda col, i: f"=IFERROR({col}75/{col}14, 0)", fmt=PCT)

    ws.freeze_panes = "C8"


# ---------------------------------------------------------------------------
# CASH FLOW TAB
# ---------------------------------------------------------------------------
def build_cashflow(wb):
    ws = wb.create_sheet("Cash flow")
    s = Sheet(ws)
    A = "Assumptions!"
    P = "'P&L'!"

    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 13
    for col in YEAR_COLS:
        ws.column_dimensions[col].width = 13
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 42

    # ---- Title ----
    s.cell(1, 1, "Cash Flow — 7-year projection (indirect method)", font=font_title)
    ws.merge_cells("A1:M1")
    s.cell(2, 1,
           "Builds operating cash from EBIT, layers in working capital and capex, "
           "then adds equity injections to derive closing cash. The minimum closing-cash "
           "point is the binding funding requirement.",
           font=font_note)
    ws.merge_cells("A2:M2")

    # ---- Year header rows ----
    s.label(4, "Year")
    for i, col in enumerate(YEAR_COLS):
        s.cell(4, 3 + i, f"={A}{col}4",
               font=font_year_label, fill=fill_year,
               align=Alignment(horizontal="center"))
    s.label(5, "Period number")
    s.year_formulas(5, lambda col, i: f"={A}{col}5", fmt=INT)
    s.label(6, "Phase")
    s.year_formulas(6, lambda col, i: f"={A}{col}7", fmt=INT)

    # ---- Section 1: Operating cash flow ----
    s.section(8, "1. OPERATING CASH FLOW")
    s.label(9, "EBIT (operating profit / loss)", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(9, 3 + i, f"={P}{col}63", font=font_link, fmt=GBP)
    s.label(10, "+ Depreciation & amortisation (non-cash)", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(10, 3 + i, f"={P}{col}61", font=font_link, fmt=GBP)
    s.label(11, "= EBITDA equivalent", units="£")
    s.year_formulas(11, lambda col, i: f"={col}9 + {col}10", fmt=GBP)
    s.label(12, "− Tax paid / + R&D credit received", units="£",
            source="Assumes credit received in year incurred (1-yr lag in reality)")
    for i, col in enumerate(YEAR_COLS):
        s.cell(12, 3 + i, f"=-{P}{col}72", font=font_formula, fmt=GBP)
    s.label(13, "= Operating cash before working capital", units="£", total=True)
    s.year_formulas(13, lambda col, i: f"={col}11 + {col}12", fmt=GBP, total=True)

    # ---- Section 2: Working capital ----
    s.section(15, "2. WORKING CAPITAL CHANGES   (positive = cash inflow)")
    s.label(16, "Inventory (end of year)", units="£",
            source="DIO × COGS / 365")
    for i, col in enumerate(YEAR_COLS):
        s.cell(16, 3 + i,
               f"={A}$B$110 * {P}{col}21 / 365",
               font=font_formula, fmt=GBP)
    s.label(17, "(Increase) / decrease in inventory", units="£")
    s.year_formulas(17,
        lambda col, i: f"=-({col}16 - OFFSET({col}16, 0, -1))", fmt=GBP)

    s.label(18, "Receivables (end of year)", units="£",
            source="DSO × revenue / 365")
    for i, col in enumerate(YEAR_COLS):
        s.cell(18, 3 + i,
               f"={A}$B$111 * {P}{col}14 / 365",
               font=font_formula, fmt=GBP)
    s.label(19, "(Increase) / decrease in receivables", units="£")
    s.year_formulas(19,
        lambda col, i: f"=-({col}18 - OFFSET({col}18, 0, -1))", fmt=GBP)

    s.label(20, "Payables (end of year)", units="£",
            source="DPO × (COGS + non-payroll opex) / 365")
    for i, col in enumerate(YEAR_COLS):
        s.cell(20, 3 + i,
               f"={A}$B$112 * ({P}{col}21 + ({P}{col}56 - {A}{col}69)) / 365",
               font=font_formula, fmt=GBP)
    s.label(21, "Increase / (decrease) in payables", units="£")
    s.year_formulas(21,
        lambda col, i: f"={col}20 - OFFSET({col}20, 0, -1)", fmt=GBP)

    s.label(22, "Deferred revenue (end of year)", units="£",
            source="≈ subscription revenue × 0.5 (annual prepay convention)")
    for i, col in enumerate(YEAR_COLS):
        s.cell(22, 3 + i,
               f"={P}{col}10 * 0.5",
               font=font_formula, fmt=GBP)
    s.label(23, "Increase / (decrease) in deferred revenue", units="£")
    s.year_formulas(23,
        lambda col, i: f"={col}22 - OFFSET({col}22, 0, -1)", fmt=GBP)

    s.label(24, "Total working capital change", units="£", total=True)
    s.year_formulas(24,
        lambda col, i: f"={col}17 + {col}19 + {col}21 + {col}23",
        fmt=GBP, total=True)

    s.label(25, "+ Interest income on opening cash", units="£",
            source="Mirrors P&L line; cash-effective in same year")
    for i, col in enumerate(YEAR_COLS):
        s.cell(25, 3 + i, f"={P}{col}64", font=font_link, fmt=GBP)

    s.label(26, "NET CASH FROM OPERATIONS", units="£", total=True)
    s.year_formulas(26, lambda col, i: f"={col}13 + {col}24 + {col}25", fmt=GBP, total=True)

    # ---- Section 3: Investing ----
    s.section(28, "3. INVESTING CASH FLOW")
    s.label(29, "Capex (cash outflow)", units="£",
            source="Linked from Assumptions capex schedule")
    for i, col in enumerate(YEAR_COLS):
        s.cell(29, 3 + i, f"=-{A}{col}103", font=font_formula, fmt=GBP)
    s.label(30, "NET CASH FROM INVESTING", units="£", total=True)
    s.year_formulas(30, lambda col, i: f"={col}29", fmt=GBP, total=True)

    # ---- Section 4: Financing ----
    s.section(32, "4. FINANCING CASH FLOW")
    s.label(33, "Equity injections", units="£",
            source="Mechanical plug — sized to keep closing cash ≥ buffer")
    for i, col in enumerate(YEAR_COLS):
        s.cell(33, 3 + i, f"={A}{col}126", font=font_link, fmt=GBP)
    s.label(34, "NET CASH FROM FINANCING", units="£", total=True)
    s.year_formulas(34, lambda col, i: f"={col}33", fmt=GBP, total=True)

    # ---- Section 5: Cash position ----
    s.section(36, "5. CASH POSITION")
    s.label(37, "Net change in cash", units="£")
    s.year_formulas(37,
        lambda col, i: f"={col}26 + {col}30 + {col}34", fmt=GBP)
    s.label(38, "Opening cash", units="£")
    s.year_formulas(38, lambda col, i: f"=OFFSET({col}38, 1, -1)", fmt=GBP)
    s.label(39, "Closing cash", units="£", total=True)
    s.year_formulas(39, lambda col, i: f"={col}38 + {col}37", fmt=GBP, total=True)

    s.label(40, "Minimum cash buffer (target floor)", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(40, 3 + i, f"={A}$B$127", font=font_link, fmt=GBP)
    s.label(41, "Buffer breached?  (1 = closing cash below buffer)", units="flag")
    s.year_formulas(41,
        lambda col, i: f"=IF({col}39 < {col}40, 1, 0)", fmt=INT)
    s.label(42, "Cumulative ops + investing  (excludes financing)", units="£",
            source="Running unfunded cash position; lowest point sets the funding floor")
    s.year_formulas(42,
        lambda col, i: f"={col}26 + {col}30 + OFFSET({col}42, 0, -1)", fmt=GBP)

    # ---- Section 6: Summary panel ----
    s.section(44, "6. CASH FLOW SUMMARY")
    s.label(45, "Minimum closing-cash point over horizon", units="£",
            source="With the equity plug — must stay above the buffer")
    s.cell(45, 2, "=MIN(C39:I39)", font=font_formula, fmt=GBP)
    s.label(46, "Year of minimum closing cash")
    s.cell(46, 2,
           '=INDEX($C$4:$K$4, MATCH(MIN(C39:I39), C39:I39, 0))',
           font=font_formula)
    s.label(47, "Peak unfunded cash trough", units="£",
            source="Most-negative cumulative ops+investing — the gross funding need")
    s.cell(47, 2, "=MIN(C42:K42)", font=font_formula, fmt=GBP)
    s.label(48, "Year of peak unfunded trough")
    s.cell(48, 2,
           '=INDEX($C$4:$K$4, MATCH(MIN(C42:K42), C42:K42, 0))',
           font=font_formula)
    s.label(49, "Total equity raised through horizon", units="£")
    s.cell(49, 2, "=SUM(C33:K33)", font=font_formula, fmt=GBP)
    s.label(50, "Cumulative net cash from operations through Y7", units="£")
    s.cell(50, 2, "=SUM(C26:K26)", font=font_formula, fmt=GBP)
    s.label(51, "Total capex through horizon", units="£")
    s.cell(51, 2, "=-SUM(C29:K29)", font=font_formula, fmt=GBP)
    s.label(52, "First year of positive operating cash flow")
    s.cell(52, 2,
           '=IFERROR(INDEX($C$4:$K$4, MATCH(TRUE, C26:K26>0, 0)), "—")',
           font=font_formula)
    s.label(53, "First year of self-funded positive cash  (ops + investing)")
    s.cell(53, 2,
           '=IFERROR(INDEX($C$4:$K$4, MATCH(TRUE, (C26:K26+C30:K30)>0, 0)), "—")',
           font=font_formula)

    # Fill the highlighted summary rows for visual block
    for r in range(45, 54):
        for col in range(2, LAST_COL + 1):
            s.ws.cell(r, col).fill = fill_subsection

    ws.freeze_panes = "C8"


# ---------------------------------------------------------------------------
# BALANCE SHEET TAB
# ---------------------------------------------------------------------------
def build_bs(wb):
    ws = wb.create_sheet("Balance sheet")
    s = Sheet(ws)
    A = "Assumptions!"
    P = "'P&L'!"
    CF = "'Cash flow'!"

    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 13
    for col in YEAR_COLS:
        ws.column_dimensions[col].width = 13
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 42

    # ---- Title ----
    s.cell(1, 1, "Balance Sheet — 7-year projection (year-end)", font=font_title)
    ws.merge_cells("A1:M1")
    s.cell(2, 1,
           "Aggregates working-capital and PP&E balances from the Cash flow / Assumptions tabs, "
           "plus cumulative share capital and retained earnings. Balance check row must be £0 every year.",
           font=font_note)
    ws.merge_cells("A2:M2")

    # ---- Year headers ----
    s.label(4, "Year")
    for i, col in enumerate(YEAR_COLS):
        s.cell(4, 3 + i, f"={A}{col}4",
               font=font_year_label, fill=fill_year,
               align=Alignment(horizontal="center"))
    s.label(5, "Period number")
    s.year_formulas(5, lambda col, i: f"={A}{col}5", fmt=INT)
    s.label(6, "Phase")
    s.year_formulas(6, lambda col, i: f"={A}{col}7", fmt=INT)

    # ---- Section 1: Assets ----
    s.section(8, "1. ASSETS")
    s.label(9, "Cash and cash equivalents", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(9, 3 + i, f"={CF}{col}39", font=font_link, fmt=GBP)
    s.label(10, "Accounts receivable", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(10, 3 + i, f"={CF}{col}18", font=font_link, fmt=GBP)
    s.label(11, "Inventory", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(11, 3 + i, f"={CF}{col}16", font=font_link, fmt=GBP)
    s.label(12, "Total current assets", units="£", total=True)
    s.year_formulas(12, lambda col, i: f"=SUM({col}9:{col}11)", fmt=GBP, total=True)

    s.label(14, "PP&E (gross — cumulative capex)", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(14, 3 + i, f"={A}{col}104", font=font_link, fmt=GBP)
    s.label(15, "Less: accumulated depreciation", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(15, 3 + i, f"=-{A}{col}106", font=font_formula, fmt=GBP)
    s.label(16, "Net PP&E", units="£", total=True)
    s.year_formulas(16, lambda col, i: f"={col}14 + {col}15", fmt=GBP, total=True)

    s.label(18, "TOTAL ASSETS", units="£", total=True)
    s.year_formulas(18, lambda col, i: f"={col}12 + {col}16", fmt=GBP, total=True)

    # ---- Section 2: Liabilities ----
    s.section(20, "2. LIABILITIES")
    s.label(21, "Accounts payable", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(21, 3 + i, f"={CF}{col}20", font=font_link, fmt=GBP)
    s.label(22, "Deferred revenue (annual prepay)", units="£")
    for i, col in enumerate(YEAR_COLS):
        s.cell(22, 3 + i, f"={CF}{col}22", font=font_link, fmt=GBP)
    s.label(23, "Total current liabilities", units="£", total=True)
    s.year_formulas(23, lambda col, i: f"={col}21 + {col}22", fmt=GBP, total=True)

    s.label(25, "TOTAL LIABILITIES", units="£", total=True)
    s.year_formulas(25, lambda col, i: f"={col}23", fmt=GBP, total=True)

    # ---- Section 3: Equity ----
    s.section(27, "3. EQUITY")
    s.label(28, "Share capital (cumulative equity raised)", units="£",
            source="Running sum of equity injections")
    for i, col in enumerate(YEAR_COLS):
        s.cell(28, 3 + i,
               f"=SUM({A}$C$126:{A}{col}126)",
               font=font_formula, fmt=GBP)
    s.label(29, "Retained earnings (cumulative NI)", units="£",
            source="Running sum of P&L net income")
    for i, col in enumerate(YEAR_COLS):
        s.cell(29, 3 + i,
               f"=SUM({P}$C$75:{P}{col}75)",
               font=font_formula, fmt=GBP)
    s.label(30, "TOTAL EQUITY", units="£", total=True)
    s.year_formulas(30, lambda col, i: f"={col}28 + {col}29", fmt=GBP, total=True)

    # ---- Section 4: Balance check ----
    s.section(32, "4. BALANCE CHECK")
    s.label(33, "Total liabilities + equity", units="£", total=True)
    s.year_formulas(33, lambda col, i: f"={col}25 + {col}30", fmt=GBP, total=True)
    s.label(34, "Difference  (assets − L − E; should be 0)", units="£")
    s.year_formulas(34, lambda col, i: f"={col}18 - {col}33", fmt=GBP)
    s.label(35, "BS balanced?  (1 = yes within £1 tolerance)")
    s.year_formulas(35, lambda col, i: f"=IF(ABS({col}34) < 1, 1, 0)", fmt=INT)

    ws.freeze_panes = "C8"


# ---------------------------------------------------------------------------
# SCENARIOS & SENSITIVITY TAB
# ---------------------------------------------------------------------------
# Note: this tab introduces multipliers that are referenced from Assumptions
# and P&L formulas. The references are by sheet name ("Scenarios!$B$XX") so
# build order doesn't matter — they're resolved at recalc time.
def build_scenarios(wb):
    ws = wb.create_sheet("Scenarios")
    s = Sheet(ws)

    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 50

    # ---- Title ----
    s.cell(1, 1, "Scenarios & Sensitivity", font=font_title)
    ws.merge_cells("A1:M1")
    s.cell(2, 1,
           "Change the scenario selector in cell B5 to flex the entire model. "
           "Active multipliers in section 3 are referenced by formulas in Assumptions and P&L. "
           "The tornado in section 5 is a linear approximation — for exact figures, toggle the selector and read the live model.",
           font=font_note)
    ws.merge_cells("A2:M2")

    # ---- Section 1: Scenario selector ----
    s.section(4, "1. SCENARIO SELECTOR")
    s.label(5, "Active scenario  (1 = Base, 2 = Downside, 3 = Upside, 4 = Reg Delay)")
    s.cell(5, 2, 1, font=font_input, fmt=INT,
           fill=PatternFill("solid", start_color="FFF2CC"))
    s.label(6, "Active scenario name")
    s.cell(6, 2,
           '=CHOOSE($B$5, "Base", "Downside", "Upside", "Reg Delay")',
           font=font_formula)

    # ---- Section 2: Scenario definition table ----
    s.section(8, "2. SCENARIO DEFINITION TABLE  (multipliers vs base + reg delay years)")
    s.cell(9, 1, "Driver", font=font_subheader)
    s.cell(9, 2, "Base",      font=font_subheader, align=Alignment(horizontal="center"))
    s.cell(9, 3, "Downside",  font=font_subheader, align=Alignment(horizontal="center"))
    s.cell(9, 4, "Upside",    font=font_subheader, align=Alignment(horizontal="center"))
    s.cell(9, 5, "Reg Delay", font=font_subheader, align=Alignment(horizontal="center"))
    s.cell(9, NOTES_COL, "Interpretation", font=font_subheader)

    drivers = [
        # (label, base, downside, upside, reg_delay, notes)
        ("Adoption ramp multiplier",            1.00, 0.50, 1.30, 1.00,
         "Slow ramp (0.5×) vs fast ramp (1.3×); delay shifts curve separately"),
        ("BOM unit cost multiplier",            1.00, 1.30, 0.85, 1.00,
         "Component price spike vs improved sourcing"),
        ("Subscription churn multiplier",       1.00, 1.50, 0.70, 1.00,
         "15% churn vs 7% churn"),
        ("Headcount / payroll multiplier",      1.00, 1.10, 0.95, 1.05,
         "Reg Delay: team runs longer, +5% cumulative payroll"),
        ("Marketing spend multiplier",          1.00, 1.20, 0.80, 1.00,
         "Less efficient acquisition vs better channels"),
        ("Regulatory & clinical multiplier",    1.00, 1.50, 0.80, 1.40,
         "Reg Delay: more NB fees, extra clinical data (+40%)"),
        ("Regulatory delay  (years shifted)",   0,    0,    0,    1,
         "Shifts commercial adoption curve right by N years"),
    ]
    for i, (lbl, base, dn, up, rd, note) in enumerate(drivers):
        r = 10 + i
        s.cell(r, 1, lbl, font=font_default)
        # integer formatting for the last row (delay years), 0.00 for multipliers
        fmt_here = INT if i == len(drivers) - 1 else '0.00'
        s.cell(r, 2, base, font=font_input, fmt=fmt_here)
        s.cell(r, 3, dn,   font=font_input, fmt=fmt_here)
        s.cell(r, 4, up,   font=font_input, fmt=fmt_here)
        s.cell(r, 5, rd,   font=font_input, fmt=fmt_here)
        s.cell(r, NOTES_COL, note, font=font_note)

    # ---- Section 3: Active multipliers (referenced by Assumptions and P&L) ----
    s.section(17, "3. ACTIVE MULTIPLIERS  (live values used by Assumptions and P&L)")
    active = [
        ("Adoption multiplier (active)",            10, '0.00'),
        ("BOM unit cost multiplier (active)",       11, '0.00'),
        ("Subscription churn multiplier (active)",  12, '0.00'),
        ("Headcount / payroll multiplier (active)", 13, '0.00'),
        ("Marketing spend multiplier (active)",     14, '0.00'),
        ("Regulatory & clinical multiplier (active)", 15, '0.00'),
        ("Regulatory delay  (years, active)",       16, INT),
    ]
    for i, (lbl, src_row, fmt) in enumerate(active):
        r = 18 + i
        s.cell(r, 1, lbl, font=font_subheader)
        s.cell(r, 2,
               f"=CHOOSE($B$5, $B${src_row}, $C${src_row}, $D${src_row}, $E${src_row})",
               font=font_formula, fmt=fmt,
               fill=fill_subsection)

    # ---- Section 4: Active scenario headline outputs ----
    s.section(26, "4. ACTIVE SCENARIO HEADLINE OUTPUTS  (live, recomputes when selector changes)")
    headlines = [
        ("Active scenario name",                              "=$B$6", FMT_TEXT),
        ("Y7 total revenue",                                  "='P&L'!I14", GBP),
        ("Y7 gross profit",                                   "='P&L'!I23", GBP),
        ("Y7 EBITDA",                                         "='P&L'!I58", GBP),
        ("Y7 net income",                                     "='P&L'!I75", GBP),
        ("Y7 closing cash",                                   "='Cash flow'!I39", GBP),
        ("Peak unfunded cash trough",                         "='Cash flow'!B47", GBP),
        ("First year of positive operating cash flow",        "='Cash flow'!B52", FMT_TEXT),
        ("Total equity raised through Y7",                    "='Cash flow'!B49", GBP),
    ]
    for i, (lbl, formula, fmt) in enumerate(headlines):
        r = 27 + i
        s.cell(r, 1, lbl, font=font_subheader)
        s.cell(r, 2, formula, font=font_link, fmt=fmt if fmt != FMT_TEXT else None,
               fill=fill_subsection)

    # ---- Section 5: Tornado sensitivity ----
    s.section(37, "5. APPROXIMATE TORNADO SENSITIVITY  (Y7 EBITDA delta vs base, linear approx)")
    s.cell(38, 1, "Driver", font=font_subheader)
    s.cell(38, 2, "Down delta", font=font_subheader, align=Alignment(horizontal="center"))
    s.cell(38, 3, "Up delta",   font=font_subheader, align=Alignment(horizontal="center"))
    s.cell(38, 4, "Magnitude",  font=font_subheader, align=Alignment(horizontal="center"))
    s.cell(38, NOTES_COL, "Approximation basis", font=font_subheader)

    # Tornado entries: (label, down_formula, up_formula, notes)
    # Refined set covering all four driver families: top-line (volume, price-hardware,
    # price-subscription, retention), cost (BOM, headcount). Reg-delay is a discrete
    # timing shock and is appended in generate_report.py from the actual scenario data.
    tornado = [
        ("Adoption ramp ±20%",
         "=-0.20 * 'P&L'!I23",
         "=0.20 * 'P&L'!I23",
         "±20% × Y7 gross profit (revenue scales linearly with adoption)"),
        ("Blended ARPU ±15%",
         "=-0.15 * 'P&L'!I10 * 0.85",
         "=0.15 * 'P&L'!I10 * 0.85",
         "±15% × Y7 sub revenue × (1 - sub COGS rate); subscription pricing lever"),
        ("BOM unit cost +30%/-15%",
         "=-0.30 * 'P&L'!I17",
         "=0.15 * 'P&L'!I17",
         "BOM change × Y7 hardware COGS (asymmetric — costs rarely fall fast)"),
        ("Subscription churn 10%→15%",
         "=-0.333 * 'P&L'!I10 * 0.85",
         "=0.333 * 'P&L'!I10 * 0.85",
         "≈ 33% sub rev reduction: steady-state active subs scale as 1/churn, so 10%→15% gives 1/1.5 = 67% of base"),
        ("Hardware ASP ±15%",
         "=-0.15 * 'P&L'!I9",
         "=0.15 * 'P&L'!I9",
         "±15% × Y7 hardware revenue (price flows directly to contribution)"),
        ("Headcount/payroll ±10%",
         "=-0.10 * Assumptions!I69",
         "=0.10 * Assumptions!I69",
         "±10% × Y7 total people cost"),
    ]
    for i, (lbl, dn, up, note) in enumerate(tornado):
        r = 39 + i
        s.cell(r, 1, lbl, font=font_default)
        s.cell(r, 2, dn,   font=font_formula, fmt=GBP)
        s.cell(r, 3, up,   font=font_formula, fmt=GBP)
        s.cell(r, 4,
               f"=MAX(ABS(B{r}), ABS(C{r}))",
               font=font_formula, fmt=GBP)
        s.cell(r, NOTES_COL, note, font=font_note)

    # Header note for the section
    s.label(46, "Read this:", subheader=True)
    s.cell(47, 1,
           "The driver with the largest magnitude is the most binding sensitivity. "
           "Adoption and BOM are typically the top two for hardware-led medtech. "
           "These are linear approximations — for second-order effects (and for the reg delay structural shift), "
           "toggle the scenario selector above.",
           font=font_note)
    ws.merge_cells("A47:M47")

    ws.freeze_panes = "C8"


# ---------------------------------------------------------------------------
# RECALC  (LibreOffice headless convert-to round-trip — populates cached values)
# ---------------------------------------------------------------------------
def recalc(path):
    import subprocess, shutil, os
    tmp_dir = "/tmp"
    subprocess.run(
        ["soffice", "--headless", "--calc", "--convert-to", "xlsx",
         "--outdir", tmp_dir, path],
        check=True, capture_output=True,
    )
    tmp_path = os.path.join(tmp_dir, os.path.basename(path))
    shutil.move(tmp_path, path)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    build_assumptions(wb)
    build_unit_economics(wb)
    build_pnl(wb)
    build_cashflow(wb)
    build_bs(wb)
    build_scenarios(wb)
    wb.save(OUT)
    recalc(OUT)
    print(f"Saved and recalculated: {OUT}")


if __name__ == "__main__":
    main()
