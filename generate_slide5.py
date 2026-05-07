"""
Generate slide 5 of the presentation: 3x2 financial overview grid.

Architecture:
    - Six individual panel PDFs (one per plot), built with the project's
      Carlito/unicode-math preamble so fonts match the report.
    - One combiner PDF that places the panels in a 3x2 grid with
      narrative-question column headers.

Layout:
    Col 1 ("What we sell, what it costs"): Revenue (top), OpEx (bottom)
    Col 2 ("Does it make money?"):         Gross margin (top), EBITDA (bottom)
    Col 3 ("Can we fund it?"):             Cash position (top), Working capital (bottom)

Reads from financial_model.xlsx (same source as generate_report.py) so numbers
stay in sync with the report.

Run:
    python3 generate_slide5.py
"""

import os
import subprocess
from openpyxl import load_workbook

MODEL    = "/home/rahul/thesis_template/financial_model.xlsx"
OUT_DIR  = "/home/rahul/thesis_template/figures/slides"
COMBINED = "slide5_financial_overview"

os.makedirs(OUT_DIR, exist_ok=True)

YEAR_LABELS = ["Y1", "Y2", "Y3", "Y4", "Y5", "Y6", "Y7"]
N = len(YEAR_LABELS)

# Project preamble — mirrors the canva-styled report figures:
#  - Carlito + Latin Modern Math (matches the report)
#  - Cream slide background RGB(246,244,241)
#  - Heavier axis/grid lines, darker plot strokes, larger marks
PANEL_PREAMBLE = r"""\documentclass[border=8pt]{standalone}
\usepackage{fontspec}
\setmainfont{Carlito}
\usepackage{unicode-math}
\setmathfont{Latin Modern Math}
\usepackage{xcolor}
\usepackage{pgfplots}
\pgfplotsset{compat=1.18}
\usetikzlibrary{positioning}

\definecolor{canvabg}{RGB}{246,244,241}
\pagecolor{canvabg}
"""

PANEL_AXIS_COMMON = r"""        width=7cm,
        height=5cm,
        symbolic x coords={Y1,Y2,Y3,Y4,Y5,Y6,Y7},
        xtick=data,
        xlabel={Year},
        grid=major,
        grid style={line width=.3pt, draw=gray!45},
        major grid style={line width=.4pt, draw=gray!55},
        axis line style={line width=0.9pt, draw=gray!70},
        tick style={line width=0.7pt, color=gray!70},
        tick label style={font=\small},
        label style={font=\normalsize},
        title style={font=\small\bfseries, yshift=-2pt},
        legend style={
            font=\scriptsize,
            cells={anchor=west},
            inner sep=2pt,
            row sep=-1pt,
            fill=canvabg,
            draw=gray!50,
        },
"""


def years(ws, row, col_start=3, n=N):
    return [ws.cell(row, col_start + i).value or 0 for i in range(n)]


def coord_k(series):
    return " ".join(f"({y},{v / 1000:.1f})" for y, v in zip(YEAR_LABELS, series))


def coord_pct(series):
    return " ".join(f"({y},{v * 100:.1f})" for y, v in zip(YEAR_LABELS, series))


def phase_markers(ymin, ymax):
    """Dashed phase dividers + small P1/P2/P3 labels at the top of the plot.
    Phase 1 = Y1-Y2 (regulatory burn), Phase 2 = Y3-Y4 (institutional bridge),
    Phase 3 = Y5-Y7 (mass-market scale)."""
    return rf"""    \draw[dashed,gray!55,line width=0.55pt] (axis cs:Y2,{ymin}) -- (axis cs:Y2,{ymax});
    \draw[dashed,gray!55,line width=0.55pt] (axis cs:Y4,{ymin}) -- (axis cs:Y4,{ymax});
    \node[anchor=north,font=\scriptsize\itshape,gray!75!black] at (axis cs:Y1,{ymax}) {{P1}};
    \node[anchor=north,font=\scriptsize\itshape,gray!75!black] at (axis cs:Y3,{ymax}) {{P2}};
    \node[anchor=north,font=\scriptsize\itshape,gray!75!black] at (axis cs:Y6,{ymax}) {{P3}};"""


def write_panel(name, axis_body):
    """Write a standalone TikZ panel and compile it to PDF."""
    tex = PANEL_PREAMBLE + "\\begin{document}\n\\begin{tikzpicture}\n" + axis_body + "\n\\end{tikzpicture}\n\\end{document}\n"
    path = os.path.join(OUT_DIR, f"{name}.tex")
    with open(path, "w") as f:
        f.write(tex)
    return path


def compile_tex(path):
    name = os.path.basename(path)
    result = subprocess.run(
        ["lualatex", "-interaction=nonstopmode", "-output-directory", OUT_DIR, name],
        cwd=OUT_DIR, capture_output=True, text=True,
    )
    pdf = path.replace(".tex", ".pdf")
    if not os.path.exists(pdf):
        print(f"FAILED: {name}")
        print("\n".join(result.stdout.splitlines()[-25:]))
        return None
    return pdf


# ---------------------------------------------------------------------------
# Load data from the financial model
# ---------------------------------------------------------------------------
print("Loading financial model...")
wb = load_workbook(MODEL, data_only=True)
PL = wb["P&L"]
CF = wb["Cash flow"]

hw_rev          = years(PL, 9)
sub_rev         = years(PL, 10)
research_rev    = years(PL, 11)
acc_rev         = years(PL, 12)
warranty_rev    = years(PL, 13)
gross_margin    = years(PL, 24)
total_rd        = years(PL, 29)
total_reg       = years(PL, 36)
total_sm        = years(PL, 43)
total_ga        = years(PL, 54)
ebitda          = years(PL, 58)
total_wc_change = years(CF, 24)
closing_cash    = years(CF, 39)
cum_unfunded    = years(CF, 42)

cum_wc_running, _running = [], 0
for v in total_wc_change:
    _running += v
    cum_wc_running.append(_running)

trough_val   = min(cum_unfunded)
trough_year  = YEAR_LABELS[cum_unfunded.index(trough_val)]
trough_label = f"\\pounds{abs(trough_val) / 1e6:.2f}M trough"
breakeven_year = next((y for y, e in zip(YEAR_LABELS, ebitda) if e > 0), "Y7")
gm_y7 = gross_margin[-1] * 100

# Revenue mix series (skip zero series)
revenue_series = [
    ("Hardware",     "blue!60",     "blue!85!black",     hw_rev),
    ("Subscription", "teal!60",     "teal!85!black",     sub_rev),
    ("Research",     "orange!60",   "orange!85!black",   research_rev),
    ("Accessories",  "purple!40",   "purple!75!black",   acc_rev),
    ("Warranty",     "gray!40",     "gray!75!black",     warranty_rev),
]
revenue_active = [(n, fc, dc, s) for (n, fc, dc, s) in revenue_series if any(v for v in s)]
revenue_addplots = "\n".join(
    rf"    \addplot+[ybar,fill={fc},draw={dc},line width=0.7pt] coordinates {{{coord_k(s)}}};"
    for (_, fc, dc, s) in revenue_active
)
revenue_legend = ", ".join(n for (n, _, _, _) in revenue_active)


# ---------------------------------------------------------------------------
# Panel 1: Revenue mix
# ---------------------------------------------------------------------------
panel_revenue = rf"""    \begin{{axis}}[
{PANEL_AXIS_COMMON}        title={{Revenue (\pounds k)}},
        ylabel={{Revenue}},
        ybar stacked,
        bar width=8pt,
        ymin=0, ymax=4700,
        legend pos=north west,
        legend columns=1,
    ]
{revenue_addplots}
    \legend{{{revenue_legend}}}
    \end{{axis}}"""


# ---------------------------------------------------------------------------
# Panel 2: Gross margin
# ---------------------------------------------------------------------------
gm_filtered_coords = " ".join(
    f"({y},{v * 100:.1f})" for y, v in zip(YEAR_LABELS, gross_margin) if v > 0
)

panel_gm = rf"""    \begin{{axis}}[
{PANEL_AXIS_COMMON}        title={{Gross margin (\%)}},
        ylabel={{GM \%}},
        ymin=30, ymax=80,
        xmin=Y1, xmax=Y7,
        xtick={{Y1,Y2,Y3,Y4,Y5,Y6,Y7}},
    ]
    \addplot[line width=1.4pt,blue!85!black,mark=*,mark size=2.5pt,mark options={{solid,fill=blue!85!black}}]
        coordinates {{{gm_filtered_coords}}};
    \end{{axis}}"""


# ---------------------------------------------------------------------------
# Panel 3: Cash position
# ---------------------------------------------------------------------------
panel_cash = rf"""    \begin{{axis}}[
{PANEL_AXIS_COMMON}        title={{Cash trough (\pounds k)}},
        ylabel={{Cash}},
        ymin=-4500, ymax=500,
    ]
    \addplot[line width=1.4pt,red!75!black,dashed,mark=square*,mark size=2.5pt,mark options={{solid,fill=red!75!black}}]
        coordinates {{{coord_k(cum_unfunded)}}};
    \draw[dotted,gray!90!black,line width=0.6pt] (axis cs:Y1,0) -- (axis cs:Y7,0);
    \node[anchor=south,font=\small\bfseries,red!75!black]
        at (axis cs:Y6,-2700) {{{trough_label}}};
    \end{{axis}}"""


# ---------------------------------------------------------------------------
# Panel 4: Operating expenses
# ---------------------------------------------------------------------------
panel_opex = rf"""    \begin{{axis}}[
{PANEL_AXIS_COMMON}        title={{Operating expenses (\pounds k)}},
        ylabel={{OpEx}},
        ybar stacked,
        bar width=8pt,
        ymin=0, ymax=2700,
        legend pos=north west,
        legend columns=1,
    ]
    \addplot+[ybar,fill=blue!60,draw=blue!85!black,line width=0.7pt] coordinates {{{coord_k(total_rd)}}};
    \addplot+[ybar,fill=teal!60,draw=teal!85!black,line width=0.7pt] coordinates {{{coord_k(total_reg)}}};
    \addplot+[ybar,fill=orange!60,draw=orange!85!black,line width=0.7pt] coordinates {{{coord_k(total_sm)}}};
    \addplot+[ybar,fill=gray!50,draw=gray!75!black,line width=0.7pt] coordinates {{{coord_k(total_ga)}}};
    \legend{{R\&D, Reg \& clinical, S\&M, G\&A}}
    \end{{axis}}"""


# ---------------------------------------------------------------------------
# Panel 5: EBITDA with phase markers + breakeven
# ---------------------------------------------------------------------------
panel_ebitda = rf"""    \begin{{axis}}[
{PANEL_AXIS_COMMON}        title={{EBITDA (\pounds k)}},
        ylabel={{EBITDA}},
        ymin=-900, ymax=1200,
    ]
    \addplot[line width=1.4pt,blue!85!black,mark=*,mark size=2.5pt,mark options={{solid,fill=blue!85!black}}]
        coordinates {{{coord_k(ebitda)}}};
    \draw[dashed,gray!70,line width=0.6pt] (axis cs:Y1,0) -- (axis cs:Y7,0);
    \node[anchor=south east,font=\small\itshape,blue!85!black]
        at (axis cs:{breakeven_year},80) {{breakeven {breakeven_year}}};
    \end{{axis}}"""


# ---------------------------------------------------------------------------
# Panel 6: Working capital (bars + cumulative line)
# ---------------------------------------------------------------------------
panel_wc = rf"""    \begin{{axis}}[
{PANEL_AXIS_COMMON}        title={{Working capital (\pounds k)}},
        ylabel={{WC inflow}},
        ymin=0, ymax=450,
        legend pos=north west,
    ]
    \addplot[ybar,bar width=10pt,fill=teal!55,draw=teal!85!black,line width=0.7pt]
        coordinates {{{coord_k(total_wc_change)}}};
    \addlegendentry{{Annual}}
    \addplot[line width=1.4pt,blue!85!black,mark=*,mark size=2.5pt,mark options={{solid,fill=blue!85!black}}]
        coordinates {{{coord_k(cum_wc_running)}}};
    \addlegendentry{{Cumulative}}
    \end{{axis}}"""


# ---------------------------------------------------------------------------
# Write and compile each panel
# ---------------------------------------------------------------------------
panels = [
    ("slide5_panel_revenue", panel_revenue),
    ("slide5_panel_gm",      panel_gm),
    ("slide5_panel_cash",    panel_cash),
    ("slide5_panel_opex",    panel_opex),
    ("slide5_panel_ebitda",  panel_ebitda),
    ("slide5_panel_wc",      panel_wc),
]

for name, body in panels:
    print(f"  building {name}...")
    write_panel(name, body)
    compile_tex(os.path.join(OUT_DIR, f"{name}.tex"))


# ---------------------------------------------------------------------------
# Combiner: 3x2 grid with narrative-question column headers
# ---------------------------------------------------------------------------
combiner_tex = r"""% Auto-generated by generate_slide5.py - do not edit manually
\documentclass[border=12pt]{standalone}
\usepackage{fontspec}
\setmainfont{Carlito}
\usepackage{xcolor}
\usepackage{graphicx}
\usepackage{tikz}
\usetikzlibrary{positioning,calc}

\definecolor{canvabg}{RGB}{246,244,241}
\pagecolor{canvabg}

\begin{document}
\begin{tikzpicture}[
    panel/.style={inner sep=0pt, outer sep=0pt, anchor=center},
    colhead/.style={font=\Large\bfseries, align=center, text=black!85, inner sep=2pt},
]
    %% Top row of panels - uniform width forces strict column alignment
    \node[panel] (rev) {\includegraphics[width=7.0cm]{slide5_panel_revenue.pdf}};
    \node[panel, right=0.6cm of rev]  (gm)   {\includegraphics[width=7.0cm]{slide5_panel_gm.pdf}};
    \node[panel, right=0.6cm of gm]   (cash) {\includegraphics[width=7.0cm]{slide5_panel_cash.pdf}};

    %% Bottom row of panels - aligned vertically against top row
    \node[panel, below=0.4cm of rev]  (opex)   {\includegraphics[width=7.0cm]{slide5_panel_opex.pdf}};
    \node[panel, below=0.4cm of gm]   (ebitda) {\includegraphics[width=7.0cm]{slide5_panel_ebitda.pdf}};
    \node[panel, below=0.4cm of cash] (wc)     {\includegraphics[width=7.0cm]{slide5_panel_wc.pdf}};

    %% Column headers - all anchored to same y line (rev.north + 14pt) so they baseline-align
    %% regardless of any panel-bbox height differences. xshift centres over the plot area.
    \node[colhead, anchor=south] at ($(rev.north)  + (12pt, 14pt)$) {Revenue and Costs};
    \node[colhead, anchor=south] at ($(rev.north -| gm)   + (12pt, 14pt)$) {Path to Profitability};
    \node[colhead, anchor=south] at ($(rev.north -| cash) + (12pt, 14pt)$) {Capital Requirement};
\end{tikzpicture}
\end{document}
"""

combiner_path = os.path.join(OUT_DIR, f"{COMBINED}.tex")
with open(combiner_path, "w") as f:
    f.write(combiner_tex)
print(f"  building combiner...")
compile_tex(combiner_path)

# ---------------------------------------------------------------------------
# Convert combined PDF to PNG for direct paste into slide deck
# ---------------------------------------------------------------------------
combined_pdf = os.path.join(OUT_DIR, f"{COMBINED}.pdf")
combined_png = os.path.join(OUT_DIR, COMBINED)  # pdftoppm appends .png with -singlefile

if subprocess.run(["which", "pdftoppm"], capture_output=True).returncode == 0:
    print("  rendering PNG...")
    subprocess.run(
        ["pdftoppm", "-png", "-r", "300", "-singlefile", combined_pdf, combined_png],
        capture_output=True, text=True,
    )
    print(f"PNG: {combined_png}.png")
else:
    print("pdftoppm not found; PNG skipped. Install with: sudo apt install poppler-utils")

print()
print(f"Done. PDF: {combined_pdf}")
